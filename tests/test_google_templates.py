import io
from datetime import date

import pandas as pd
from openpyxl import Workbook

from app.exports.google_templates import (
    _sign_in_client_rows,
    _sign_in_employee_rows,
    build_sign_in_sheet_pdf,
    build_pdf_image_print_html,
    get_google_template_workbook_bytes,
    load_template_sheet_workbook,
)


def _template_bytes() -> bytes:
    wb = Workbook()
    wb.active.title = "Daily Time"
    wb.create_sheet("TimeEntries")
    sign_in = wb.create_sheet("Sign In Sheet")
    sign_in["D6"] = ""
    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()
    buffer.seek(0)
    return buffer.getvalue()


def test_load_template_sheet_workbook_keeps_requested_google_tab_only():
    wb, ws = load_template_sheet_workbook(_template_bytes(), ("TimeEntries", "Time Entries"))

    assert ws.title == "TimeEntries"
    assert wb.sheetnames == ["TimeEntries"]


def test_google_template_export_falls_back_to_existing_session(monkeypatch):
    class FakeResponse:
        content = b"xlsx-bytes"

        def raise_for_status(self):
            return None

    class FakeSession:
        def get(self, url, params):
            assert url == "https://www.googleapis.com/drive/v3/files/sheet-id/export"
            assert params["mimeType"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            return FakeResponse()

    class FakeManager:
        def _ensure_session(self):
            return FakeSession()

    monkeypatch.setattr("app.exports.google_templates.get_sheets_manager", lambda: FakeManager())

    assert get_google_template_workbook_bytes("sheet-id") == b"xlsx-bytes"


def test_sign_in_employee_rows_map_active_columns_l_m_n():
    employees = pd.DataFrame(
        [
            ["EMPL", "S-MIL10", "ADAM MILLER", "", "", "", "", "", "", "", "", "PTW", "Supervisor", "TRUE"],
            ["EMPL", "72454", "ANDY LYNDS", "", "", "", "", "", "", "", "", "PTW", "Electrician", "FALSE"],
            ["EMPL", "17446", "TRAVIS TYCHKOWSKY", "", "", "", "", "", "", "", "", "PTW", "Welder", "Y"],
        ],
        columns=[
            "Time Record Type",
            "Person Number",
            "Employee Name",
            "Indirect / Direct",
            "Override Trade Class",
            "Truck",
            "Post To Payroll",
            "Night Shift",
            "Premium Rate",
            "Subsistence Rate",
            "Travel Rate",
            "Company Name",
            "Craft / Certification",
            "Active",
        ],
    )

    rows, active_count, first_hidden_row = _sign_in_employee_rows(employees)

    assert active_count == 2
    assert first_hidden_row == 18
    assert len(rows) == 64
    assert rows[0] == ["PTW", "ADAM MILLER", "Supervisor"]
    assert rows[1] == ["PTW", "TRAVIS TYCHKOWSKY", "Welder"]
    assert rows[2] == ["", "", ""]


def test_sign_in_client_rows_map_active_rows_and_three_blanks():
    clients = pd.DataFrame(
        [
            ["PEMBINA", "MARK SOMERS", "CONSTRUCTION MANAGER", "TRUE"],
            ["PEMBINA", "INACTIVE CLIENT", "VISITOR", "FALSE"],
            ["PEMBINA", "SCOTT RADTKE", "E&I SUPERVISOR", "Y"],
        ],
        columns=["COMPANY", "PERSON NAME", "CERTIFICATION", "Active"],
    )

    rows, active_count, first_hidden_row = _sign_in_client_rows(clients)

    assert active_count == 2
    assert first_hidden_row == 81
    assert len(rows) == 18
    assert rows[0] == ["PEMBINA", "MARK SOMERS", "CONSTRUCTION MANAGER"]
    assert rows[1] == ["PEMBINA", "SCOTT RADTKE", "E&I SUPERVISOR"]
    assert rows[2] == ["", "", ""]


def test_sign_in_sheet_pdf_batches_google_updates(monkeypatch):
    import fitz

    def pdf_bytes(label: str) -> bytes:
        document = fitz.open()
        page = document.new_page(width=612, height=792)
        page.insert_text((72, 72), label)
        content = document.tobytes()
        document.close()
        return content

    class FakeManager:
        def __init__(self):
            self.batch_update_calls = []
            self.batch_value_calls = []
            self.single_value_calls = []
            self.exported_sheet_ids = []
            self.next_sheet_id = 200

        def get_spreadsheet_metadata(self, spreadsheet_id, fields=None):
            assert spreadsheet_id == "sheet-id"
            return {
                "sheets": [
                    {"properties": {"title": "Sign In Sheet", "sheetId": 100}},
                ]
            }

        def batch_update(self, spreadsheet_id, requests_body):
            self.batch_update_calls.append(requests_body)
            if requests_body and "duplicateSheet" in requests_body[0]:
                replies = []
                for request in requests_body:
                    self.next_sheet_id += 1
                    replies.append({
                        "duplicateSheet": {
                            "properties": {
                                "sheetId": self.next_sheet_id,
                                "title": request["duplicateSheet"]["newSheetName"],
                            }
                        }
                    })
                return {"replies": replies}
            return {}

        def batch_update_values(self, spreadsheet_id, data, value_input_option="USER_ENTERED"):
            self.batch_value_calls.append(data)
            return {}

        def update_values(self, spreadsheet_id, range_name, values, value_input_option="USER_ENTERED"):
            self.single_value_calls.append((range_name, values))
            return {}

        def export_sheet_pdf(self, spreadsheet_id, sheet_id, *, repeat_frozen_rows=False, margins=None):
            self.exported_sheet_ids.append((sheet_id, repeat_frozen_rows, margins))
            return pdf_bytes(f"sheet {sheet_id}")

    fake_manager = FakeManager()
    monkeypatch.setattr("app.exports.google_templates.get_sheets_manager", lambda: fake_manager)
    monkeypatch.setattr("app.exports.google_templates.time.sleep", lambda _: None)

    employees = pd.DataFrame(
        [["PTW", "ADAM MILLER", "Supervisor", "TRUE"]],
        columns=["Company Name", "Employee Name", "Craft / Certification", "Active"],
    )
    clients = pd.DataFrame(
        [["PEMBINA", "MARK SOMERS", "SAFETY", "TRUE"]],
        columns=["COMPANY", "PERSON NAME", "CERTIFICATION", "Active"],
    )

    pdf, employee_count, client_count, sheet_count = build_sign_in_sheet_pdf(
        employees,
        [date(2026, 8, 1), date(2026, 8, 2)],
        clients,
        "sheet-id",
    )

    assert pdf.startswith(b"%PDF")
    assert employee_count == 1
    assert client_count == 1
    assert sheet_count == 2
    assert len(fake_manager.batch_update_calls[0]) == 2
    formatting_requests = fake_manager.batch_update_calls[1]
    repeat_header_requests = [
        request for request in formatting_requests
        if "updateSheetProperties" in request
    ]
    assert len(repeat_header_requests) == 2
    assert repeat_header_requests[0]["updateSheetProperties"]["properties"]["gridProperties"]["frozenRowCount"] == 10
    assert len(fake_manager.batch_value_calls) == 1
    assert len(fake_manager.batch_value_calls[0]) == 6
    assert fake_manager.single_value_calls == []
    assert fake_manager.exported_sheet_ids == [
        (201, True, {"top": 0.25, "bottom": 0.25, "left": 0.25, "right": 0.25}),
        (202, True, {"top": 0.25, "bottom": 0.25, "left": 0.25, "right": 0.25}),
    ]


def test_pdf_image_print_html_renders_images_and_calls_print():
    import fitz

    document = fitz.open()
    page = document.new_page(width=612, height=792)
    page.insert_text((72, 72), "Sign In Sheet")
    pdf_bytes = document.tobytes()
    document.close()

    html = build_pdf_image_print_html(pdf_bytes, auto_print=True)

    assert "data:image/png;base64," in html
    assert "window.print()" in html
    assert "application/pdf" not in html
    assert "Open PDF" not in html
