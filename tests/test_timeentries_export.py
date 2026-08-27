import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side

from app.exports.timeentries_export import (
    apply_daily_import_data_row_style,
    build_daily_import_rate_cells,
    filter_daily_import_rows,
)


def test_subsistence_rate_only_goes_on_subsistence_line():
    regular_rate, subsistence_rate = build_daily_import_rate_cells(
        night_shift="",
        premium_rate="",
        subsistence_rate="225",
        travel_rate="",
    )

    assert regular_rate == ""
    assert subsistence_rate == "225"


def test_regular_rate_uses_premium_or_travel_not_subsistence():
    premium_regular_rate, premium_subsistence_rate = build_daily_import_rate_cells(
        night_shift="",
        premium_rate="PREM",
        subsistence_rate="225",
        travel_rate="TRAVEL",
    )
    travel_regular_rate, travel_subsistence_rate = build_daily_import_rate_cells(
        night_shift="",
        premium_rate="",
        subsistence_rate="225",
        travel_rate="TRAVEL",
    )

    assert premium_regular_rate == "PREM"
    assert premium_subsistence_rate == "225"
    assert travel_regular_rate == "TRAVEL"
    assert travel_subsistence_rate == "225"


def test_night_shift_keeps_ns_only_on_regular_rows():
    regular_rate, subsistence_rate = build_daily_import_rate_cells(
        night_shift="Y",
        premium_rate="PREM",
        subsistence_rate="225",
        travel_rate="TRAVEL",
    )

    assert regular_rate == "NS"
    assert subsistence_rate == "225"


def test_daily_import_style_matches_template_data_row_after_row_34():
    wb = Workbook()
    ws = wb.active
    light_side = Side(style="thin", color="D9D9D9")
    black_side = Side(style="thick", color="000000")

    for col_idx in range(1, 16):
        template_cell = ws.cell(row=4, column=col_idx)
        template_cell.alignment = Alignment(horizontal="center", vertical="center")
        template_cell.border = Border(
            left=light_side,
            right=light_side,
            top=light_side,
            bottom=light_side,
        )
        ws.cell(row=33, column=col_idx).border = Border(bottom=black_side)
        ws.cell(row=34, column=col_idx).alignment = Alignment(horizontal="left", vertical="bottom")
        ws.cell(row=34, column=col_idx).border = Border(top=black_side)

    ws.row_dimensions[4].height = 18

    apply_daily_import_data_row_style(ws, 33)
    apply_daily_import_data_row_style(ws, 34)

    assert ws.cell(row=34, column=3).alignment.horizontal == "center"
    assert ws.cell(row=34, column=3).alignment.vertical == "center"
    assert ws.cell(row=33, column=3).border.bottom.style == "thin"
    assert ws.cell(row=34, column=3).border.top.style == "thin"
    assert ws.row_dimensions[34].height == 18


def test_filter_daily_import_rows_uses_employee_list_y_flag_only():
    day_df = pd.DataFrame(
        {
            "Name": ["ADAM MILLER", "TRAVIS TYCHKOWSKY", "GREGG MORRISON"],
            "Job Number": ["2624138043", "2624138043", "2624138043"],
        }
    )
    employees_df = pd.DataFrame(
        {
            "Employee Name": ["Adam Miller", "Travis Tychkowsky", "Gregg Morrison"],
            "Daily Import": ["Y", "", "TRUE"],
            "Active": ["TRUE", "TRUE", "TRUE"],
        }
    )

    filtered = filter_daily_import_rows(day_df, employees_df)

    assert filtered["Name"].tolist() == ["ADAM MILLER", "GREGG MORRISON"]


def test_filter_daily_import_rows_keeps_old_behavior_when_column_missing():
    day_df = pd.DataFrame({"Name": ["ADAM MILLER", "TRAVIS TYCHKOWSKY"]})
    employees_df = pd.DataFrame({"Employee Name": ["Adam Miller", "Travis Tychkowsky"]})

    filtered = filter_daily_import_rows(day_df, employees_df)

    assert filtered["Name"].tolist() == ["ADAM MILLER", "TRAVIS TYCHKOWSKY"]
