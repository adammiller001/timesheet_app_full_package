# --- begin export_daily_time.py (V13) ---
from __future__ import annotations
from collections import defaultdict
from datetime import date as date_cls
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

try:
    from app.config import script_dir
except Exception:
    def script_dir() -> Path:
        return Path(__file__).resolve().parents[2]

DAILY_TEMPLATE = "Daily Time.xlsx"
TIMEENTRIES_TEMPLATE = "TimeEntries.xlsx"
EMPLOYEE_BOOK = "Timesheet Apps.xlsx"
EMPLOYEE_SHEET = "Employee List"
PAYCODES_SHEET = "Pay Codes"

def _norm(s: str) -> str:
    return " ".join(str(s).strip().lower().split())

def _resolve_in_root(*names: str) -> Optional[Path]:
    root = script_dir()
    for name in names:
        p = root / name
        if p.exists():
            return p
    return None

def _parse_job(job: str) -> Tuple[str, str, str]:
    parts = [p.strip() for p in str(job).split(" - ")]
    if len(parts) >= 3: return parts[0], parts[1], " - ".join(parts[2:])
    if len(parts) == 2: return parts[0], parts[1], ""
    if parts:           return parts[0], "", ""
    return "", "", ""

def _fmt_job_full(job: str) -> str:
    j, a, d = _parse_job(job)
    a3 = a.zfill(3) if a else ""
    if j and a3 and d: return f"{j} - {a3} - {d}"
    if j and a3:       return f"{j} - {a3}"
    return job

def _parse_cost(cost: str) -> Tuple[str, str]:
    parts = [p.strip() for p in str(cost).split(" - ", 1)]
    if len(parts) == 2: return parts[0], parts[1]
    if parts:           return parts[0], ""
    return "", ""

# ---------- Employee lookups (A/E/F/G), classification, pay codes ----------
from io import BytesIO

def _load_employee_sheet() -> Optional[pd.DataFrame]:
    path = _resolve_in_root(EMPLOYEE_BOOK)
    if not path: return None
    raw = path.read_bytes()
    xl = pd.ExcelFile(BytesIO(raw))
    try:
        return xl.parse(EMPLOYEE_SHEET)
    except Exception:
        try:
            return xl.parse(0)
        except Exception:
            return None

EMP_NAME_COLS = ["Employee Name", "Employee", "Name"]
EMP_ID_COLS   = ["Person Number", "Employee #", "Employee Number", "Person #", "Emp #", "Emp Number", "ID", "EmpID"]

def _find_col(df: pd.DataFrame, candidates) -> Optional[str]:
    lut = { _norm(c): c for c in df.columns }
    for c in candidates:
        k = _norm(c)
        if k in lut: return lut[k]
    return None

def _emp_maps():
    df = _load_employee_sheet()
    if df is None or df.empty:
        return {}, {}, {}, {}, {}
    name_col = _find_col(df, EMP_NAME_COLS) or df.columns[0]
    id_col   = _find_col(df, EMP_ID_COLS)

    colA = df.columns[0] if len(df.columns) > 0 else None
    colE = df.columns[4] if len(df.columns) > 4 else None
    colF = df.columns[5] if len(df.columns) > 5 else None
    colG = df.columns[6] if len(df.columns) > 6 else None

    by_idA, by_nameA = {}, {}
    by_idE, by_nameE = {}, {}
    by_idF, by_nameF = {}, {}
    by_idG, by_nameG = {}, {}

    for _, row in df.iterrows():
        nm  = _norm(row.get(name_col, ""))
        pid = str(row.get(id_col, "")).strip() if id_col else ""
        A_val = row.get(colA, "") if colA else ""
        E_val = row.get(colE, "") if colE else ""
        F_val = row.get(colF, "") if colF else ""
        G_val = row.get(colG, "") if colG else ""
        if pid:
            by_idA[_norm(pid)] = A_val; by_idE[_norm(pid)] = E_val
            by_idF[_norm(pid)] = F_val; by_idG[_norm(pid)] = G_val
        if nm:
            by_nameA[nm] = A_val; by_nameE[nm] = E_val
            by_nameF[nm] = F_val; by_nameG[nm] = G_val
    return (by_idA, by_nameA, by_idE, by_nameE, by_idF, by_nameF, by_idG, by_nameG)

def _load_paycodes() -> Dict[str, str]:
    path = _resolve_in_root(EMPLOYEE_BOOK)
    res: Dict[str,str] = {}
    if not path: return res
    raw = path.read_bytes()
    xl = pd.ExcelFile(BytesIO(raw))
    try:
        df = xl.parse(PAYCODES_SHEET)
    except Exception:
        return res
    # tolerant extraction: label->code
    label_col = None; code_col = None
    for c in df.columns:
        cn = _norm(c)
        if label_col is None and any(k in cn for k in ["label","type","name","pay","earning","code type"]):
            label_col = c
        if code_col is None and "code" in cn:
            code_col = c
    if label_col and code_col:
        for _, row in df.iterrows():
            lbl  = _norm(row.get(label_col,""))
            code = str(row.get(code_col,"")).strip()
            if not code: continue
            if "rt" in lbl or "regular" in lbl:       res["rt"] = code
            if "ot" in lbl or "overtime" in lbl:      res["ot"] = code
            if "subsist" in lbl or "subsis" in lbl:   res["subsistence"] = code
        return res
    # fallback: scan cells
    h, w = df.shape
    for r in range(h):
        for c in range(w):
            try: val = _norm(df.iat[r,c])
            except Exception: continue
            if not val: continue
            if any(k in val for k in ["rt","regular","ot","overtime","subsist","subsis"]):
                if c+1 < w:
                    code = str(df.iat[r,c+1]).strip()
                    if not code: continue
                    if "rt" in val or "regular" in val:     res.setdefault("rt", code)
                    if "ot" in val or "overtime" in val:    res.setdefault("ot", code)
                    if "subsist" in val or "subsis" in val: res.setdefault("subsistence", code)
    return res

# -------------- MAIN EXPORTS --------------

def export_daily_time(chosen_date, entered_df: pd.DataFrame):
    """
    Returns (daily_path, [timeentries_paths...]).
    Streamlit can then show a download button for each.
    """
    if isinstance(chosen_date, pd.Timestamp):
        chosen_date = chosen_date.date()
    assert isinstance(chosen_date, date_cls)

    daily_path = _export_daily_time_workbook(chosen_date, entered_df)
    job_paths  = _export_timeentries_files(chosen_date, entered_df)
    return daily_path, job_paths

def _export_daily_time_workbook(chosen_date, entered_df: pd.DataFrame) -> Path:
    root = script_dir()
    tmpl = _resolve_in_root(DAILY_TEMPLATE, "daily time.xlsx", "DailyTime.xlsx")
    if not tmpl:
        raise FileNotFoundError("Daily Time template not found in project root.")
    out = root / (chosen_date.strftime("%m-%d-%Y") + " - Daily Time.xlsx")

    wb = load_workbook(filename=str(tmpl))
    ws = wb.active
    ws["AA1"].value = "PTW MAP V13"
    ws["G5"].value  = chosen_date.strftime("%Y-%m-%d")

    df = entered_df.copy()
    for col in ["Employee","Person Number","Trade Class","Premium Rate","Job","Cost Code","RT Hours","OT Hours","Notes"]:
        if col not in df.columns: df[col] = ""

    # Minimal write (same layout as previous build)
    def _fmt_job_g(job: str) -> str:
        j, a, d = _parse_job(job)
        a3 = a.zfill(3) if a else ""
        if j and a3 and d: return f"{j} - {a3} - {d}"
        if j and a3:       return f"{j} - {a3}"
        return job

    def _parse_cost(cost: str) -> Tuple[str, str]:
        parts = [p.strip() for p in str(cost).split(" - ", 1)]
        if len(parts)==2: return parts[0], parts[1]
        if parts:         return parts[0], ""
        return "", ""

    row_cursor = 32  # write into the direct band to keep simple (your previous formatting remains)
    for _, r in df.iterrows():
        code, desc = _parse_cost(r.get("Cost Code",""))
        rt = float(r.get("RT Hours",0) or 0); ot = float(r.get("OT Hours",0) or 0)
        prim = [
            r.get("Employee",""), r.get("Trade Class",""), "", r.get("Premium Rate",""),
            desc, code, _fmt_job_g(r.get("Job","")), rt, ot, rt+ot
        ]
        for i, v in enumerate(prim, start=1):
            ws.cell(row=row_cursor, column=i, value=v)
        row_cursor += 1

    # Notes section (A264+), bold+underline header per job then notes then blank row
    header_font = Font(bold=True, underline="single")
    job_notes: Dict[str,set] = defaultdict(set)
    for _, r in df.iterrows():
        j = str(r.get("Job","")).strip(); n = str(r.get("Notes","")).strip()
        if j and n: job_notes[j].add(n)
    rr = 264
    for job in sorted(job_notes.keys()):
        cell = ws.cell(row=rr, column=1, value=_fmt_job_g(job)); cell.font = header_font; rr += 1
        for n in sorted(job_notes[job]): ws.cell(row=rr, column=1, value=n); rr += 1
        rr += 1

    wb.save(str(out))
    return out

def _export_timeentries_files(chosen_date, df: pd.DataFrame) -> List[Path]:
    """Create one TimeEntries export per job and return the list of paths."""
    root = script_dir()
    tmpl = _resolve_in_root(TIMEENTRIES_TEMPLATE)
    if not tmpl:
        return []

    (by_idA, by_nameA, by_idE, by_nameE, by_idF, by_nameF, by_idG, by_nameG) = _emp_maps()
    paycodes = _load_paycodes()

    def empA(pid, name): return by_idA.get(_norm(pid), None) or by_nameA.get(_norm(name), "")
    def empE(pid, name): return by_idE.get(_norm(pid), None) or by_nameE.get(_norm(name), "")
    def empF(pid, name):
        v = by_idF.get(_norm(pid), None) or by_nameF.get(_norm(name), "")
        return "Y" if str(v).strip().lower() in {"y","yes","true","1"} else ""
    def empG(pid, name): return by_idG.get(_norm(pid), None) or by_nameG.get(_norm(name), "")

    jobs = defaultdict(list)
    for _, r in df.iterrows():
        jobs[str(r.get("Job","")).strip()].append(r.to_dict())

    paths: List[Path] = []
    for job_key, rows in jobs.items():
        if not job_key: continue
        jobnum, area, _ = _parse_job(job_key); area3 = area.zfill(3) if area else ""
        wb2 = load_workbook(filename=str(tmpl)); ws2 = wb2.active
        r0 = 4
        for r in rows:
            date_str = chosen_date.strftime("%Y-%m-%d")
            emp  = r.get("Employee",""); pid = str(r.get("Person Number","")).strip()
            cost_code, _ = _parse_cost(r.get("Cost Code",""))
            # RT
            rt = float(r.get("RT Hours",0) or 0)
            if rt > 0:
                ws2.cell(row=r0, column=1, value=date_str)
                ws2.cell(row=r0, column=2, value=empA(pid, emp))
                ws2.cell(row=r0, column=3, value=pid)
                ws2.cell(row=r0, column=4, value=emp)
                ws2.cell(row=r0, column=5, value=empE(pid, emp))
                ws2.cell(row=r0, column=6, value="Y")
                ws2.cell(row=r0, column=7, value=cost_code)
                ws2.cell(row=r0, column=8, value=area3)
                ws2.cell(row=r0, column=10, value=paycodes.get("rt",""))
                ws2.cell(row=r0, column=11, value=rt)
                ws2.cell(row=r0, column=12, value=empF(pid, emp))
                ws2.cell(row=r0, column=13, value=empG(pid, emp))
                r0 += 1
            # OT
            ot = float(r.get("OT Hours",0) or 0)
            if ot > 0:
                ws2.cell(row=r0, column=1, value=date_str)
                ws2.cell(row=r0, column=2, value=empA(pid, emp))
                ws2.cell(row=r0, column=3, value=pid)
                ws2.cell(row=r0, column=4, value=emp)
                ws2.cell(row=r0, column=5, value=empE(pid, emp))
                ws2.cell(row=r0, column=6, value="Y")
                ws2.cell(row=r0, column=7, value=cost_code)
                ws2.cell(row=r0, column=8, value=area3)
                ws2.cell(row=r0, column=10, value=paycodes.get("ot",""))
                ws2.cell(row=r0, column=11, value=ot)
                ws2.cell(row=r0, column=12, value=empF(pid, emp))
                ws2.cell(row=r0, column=13, value=empG(pid, emp))
                r0 += 1
            # Subsistence from Premium text
            prem = str(r.get("Premium Rate","")).lower()
            if prem and "subsist" in prem:
                ws2.cell(row=r0, column=1, value=date_str)
                ws2.cell(row=r0, column=2, value=empA(pid, emp))
                ws2.cell(row=r0, column=3, value=pid)
                ws2.cell(row=r0, column=4, value=emp)
                ws2.cell(row=r0, column=5, value=empE(pid, emp))
                ws2.cell(row=r0, column=6, value="Y")
                ws2.cell(row=r0, column=7, value=cost_code)
                ws2.cell(row=r0, column=8, value=area3)
                ws2.cell(row=r0, column=10, value=paycodes.get("subsistence",""))
                ws2.cell(row=r0, column=11, value=1)
                ws2.cell(row=r0, column=12, value=empF(pid, emp))
                ws2.cell(row=r0, column=13, value=empG(pid, emp))
                r0 += 1

        out2 = root / f"{chosen_date.strftime('%m-%d-%Y')} - {jobnum} - Daily Import.xlsx"
        wb2.save(str(out2))
        paths.append(out2)

    return paths
# --- end export_daily_time.py ---
