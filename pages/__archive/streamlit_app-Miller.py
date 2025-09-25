
from __future__ import annotations
from io import BytesIO
from pathlib import Path
from time import sleep
from typing import Set

import pandas as pd
import streamlit as st

APP_TITLE = "PTW - Daily Timesheet Suite"
LOGO_FILE = "PTW.jpg"                   # place next to this file
EXCEL_FILENAME = "Timesheet Apps.xlsx"  # place next to this file
USERS_SHEET_CANDIDATES = ["Users", "User List", "Access", 0]
USER_EMAIL_COLS = [
    "User's Email Address",
    "Email",
    "Email Address",
    "User Email",
    "Login Email",
]

def project_root() -> Path:
    try:
        return Path(__file__).resolve().parent
    except NameError:
        return Path.cwd()

def excel_path() -> Path:
    return project_root() / EXCEL_FILENAME

def logo_path() -> Path:
    return project_root() / LOGO_FILE

def _parse_users_from_excelfile(xl: pd.ExcelFile) -> Set[str]:
    for sheet in USERS_SHEET_CANDIDATES:
        try:
            df = xl.parse(sheet)
        except Exception:
            continue
        for col in USER_EMAIL_COLS:
            if col in df.columns:
                vals = (
                    df[col]
                    .astype(str)
                    .str.strip()
                    .str.lower()
                    .replace({"nan": ""})
                    .tolist()
                )
                return {v for v in vals if v}
    raise KeyError(
        f"No Users sheet/column found. Tried sheets {USERS_SHEET_CANDIDATES} and columns {USER_EMAIL_COLS}."
    )

@st.cache_data(show_spinner=False)
def load_allowed_emails() -> Set[str]:
    """
    Robust loader with retries:
      1) Retry (8 attempts, exponential backoff) reading raw bytes and parsing via BytesIO.
      2) If PermissionError persists, try openpyxl.read_only=true (can bypass some locks).
    """
    xlsx = excel_path()
    if not xlsx.exists():
        raise FileNotFoundError(f"'{EXCEL_FILENAME}' not found next to this app at: {xlsx}")

    last_err = None
    # Strategy 1: read bytes + pandas ExcelFile
    for attempt in range(8):
        try:
            raw = xlsx.read_bytes()
            xl = pd.ExcelFile(BytesIO(raw))
            return _parse_users_from_excelfile(xl)
        except PermissionError as e:
            last_err = e
            sleep(0.2 * (2**attempt))  # backoff
        except Exception as e:
            last_err = e
            break

    # Strategy 2: openpyxl read_only
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filename=str(xlsx), read_only=True, data_only=True)
        # Pick first matching sheet
        ws = None
        for s in USERS_SHEET_CANDIDATES:
            if isinstance(s, int):
                try:
                    ws = wb.worksheets[s]
                    break
                except Exception:
                    continue
            if s in wb.sheetnames:
                ws = wb[s]
                break
        if ws is None:
            ws = wb.worksheets[0]

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Users sheet is empty.")
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        # find column
        col_idx = None
        for cname in USER_EMAIL_COLS:
            try:
                col_idx = headers.index(cname)
                break
            except ValueError:
                continue
        if col_idx is None:
            raise KeyError(
                f"No email column found. Expected one of: {USER_EMAIL_COLS}. Got headers: {headers}"
            )
        emails = set()
        for r in rows[1:]:
            if r is None or len(r) <= col_idx:
                continue
            v = r[col_idx]
            if v:
                emails.add(str(v).strip().lower())
        if not emails:
            raise ValueError("No emails found in Users sheet.")
        return emails
    except Exception as e:
        raise PermissionError(f"Could not read '{EXCEL_FILENAME}'. Last error: {last_err or e}") from e

def _hide_sidebar_pages_until_login():
    # Hide Streamlit's auto-generated page navigation in the sidebar
    st.markdown(
        """
        <style>
        [data-testid="stSidebarNav"] { display: none !important; }
        section[data-testid="stSidebar"] nav { display: none !important; }
        </style>
        """,
        unsafe_allow_html=True,
    )

def center_logo():
    lp = logo_path()
    if lp.exists():
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            # Doubled size (was 220)
            st.image(str(lp), width=440, use_container_width=False)
    else:
        st.caption(f"(Logo file '{LOGO_FILE}' not found.)")

# ---------- UI ----------
st.set_page_config(page_title=APP_TITLE, layout="wide")

# Hide sidebar pages if not signed in
if "user_email" not in st.session_state:
    _hide_sidebar_pages_until_login()

center_logo()
st.title("PTW - Daily Timesheet Suite")
st.write("Welcome! Please sign in with your **work email** to continue.")

# Load access list
try:
    emails_ok = load_allowed_emails()
except Exception as e:
    st.error(
        f"Could not load the Users list from **{EXCEL_FILENAME}**.\n\n"
        f"**Details:** {e}\n\n"
        "- Close the workbook in Excel and disable Explorer **Preview/Details** pane.\n"
        "- Pause OneDrive sync briefly.\n"
        "- Ensure the file is *Always keep on this device*.\n"
        "- File path I tried: `" + str(excel_path()) + "`"
    )
    st.stop()

# Login form
with st.form("login_form", clear_on_submit=False):
    email = st.text_input("Email address", placeholder="you@ptwenergy.com").strip().lower()
    submitted = st.form_submit_button("Enter")

if submitted:
    if not email:
        st.warning("Please enter your email address.")
    elif email not in emails_ok:
        st.error("Access denied. Your email is not on the **Users** list in the workbook.")
        st.caption("If this is unexpected, ask an admin to add your email to the Users sheet.")
    else:
        st.session_state["user_email"] = email
        st.success(f"Welcome, {email}!")
        # Optional: auto-redirect to entry page when supported
        if hasattr(st, "switch_page"):
            st.switch_page("pages/10_Timesheet_Entry.py")
        else:
            st.markdown("---")
            st.subheader("Continue")
            if hasattr(st, "page_link"):
                st.page_link("pages/10_Timesheet_Entry.py", label="Ã¢â‚¬Â¢Ã‹Å“ Timesheet Entry", icon="Ã¢â‚¬Â¢Ã‹Å“")
                st.page_link("pages/20_Whats_Been_Added_Today.py", label=""Ã‚Â What's Been Added Today", icon=""Ã‚Â")
                st.page_link("pages/30_Export_Day.py", label=""Ã‚Â¤ Export Day", icon=""Ã‚Â¤")
                st.page_link("pages/90_Admin.py", label="ÃƒÂ¢Ã…Â¡Ã¢â€žÂ¢ÃƒÂ¯Ã‚Â¸Ã‚Â Admin", icon="ÃƒÂ¢Ã…Â¡Ã¢â€žÂ¢ÃƒÂ¯Ã‚Â¸Ã‚Â")
            else:
                st.info("Use the left sidebar to open a page (Timesheet Entry, Exports, Admin).")
            st.markdown("---")
            st.caption("You can change users by refreshing the page.")
else:
    st.caption("Your email will be checked against the **Users** sheet in the workbook.")


