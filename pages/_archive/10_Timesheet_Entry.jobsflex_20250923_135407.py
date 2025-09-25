
import streamlit as st
# ===== Autoload job & cost options (robust) - injected =====
from pathlib import Path as __P
import pandas as __pd
import streamlit as __st

# ensure variables exist even if load fails
job_options  = [] if "job_options"  not in locals() else job_options
cost_options = [] if "cost_options" not in locals() else cost_options

def __normalize_cols(df):
    # map lower/stripped -> original
    m = {}
    for c in df.columns:
        if isinstance(c, str):
            k = c.strip().lower().replace("–","-").replace("—","-")
            m[k] = c
    return m

def __excel_sheets(xlsx):
    try:
        xl = __pd.ExcelFile(xlsx)
        return xl, xl.sheet_names
    except Exception:
        return None, []

def __try_read_jobs(xlsx):
    # candidate sheet names and column names (case-insensitive)
    sheets = ["jobs", "job list", "jobs list", "jobslist"]
    need   = ["job number", "area number", "description"]
    # 1) pandas via raw bytes
    try:
        raw = xlsx.read_bytes()
        xl  = __pd.ExcelFile(raw)
        for s in sheets:
            if s in [n.lower() for n in xl.sheet_names]:
                real = next(n for n in xl.sheet_names if n.lower()==s)
                df = xl.parse(real)
                cols = __normalize_cols(df)
                if all(n in cols for n in need):
                    J, A, D = (cols[n] for n in need)
                    tmp = df[[J, A, D]].astype(str).fillna("")
                    return (tmp[J] + " - " + tmp[A] + " - " + tmp[D]).tolist(), None
    except Exception as e:
        err = f"[jobs raw] {e}"
    # 2) pandas normal
    try:
        xl, names = __excel_sheets(xlsx)
        if xl:
            for s in sheets:
                if s in [n.lower() for n in names]:
                    real = next(n for n in names if n.lower()==s)
                    df = xl.parse(real)
                    cols = __normalize_cols(df)
                    if all(n in cols for n in need):
                        J, A, D = (cols[n] for n in need)
                        tmp = df[[J, A, D]].astype(str).fillna("")
                        return (tmp[J] + " - " + tmp[A] + " - " + tmp[D]).tolist(), None
    except Exception as e:
        err = f"[jobs pandas] {e}"
    # 3) openpyxl read-only
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filename=str(xlsx), read_only=True, data_only=True)
        names = [ws.title for ws in wb.worksheets]
        for s in sheets:
            if s in [n.strip().lower() for n in names]:
                real = next(n for n in names if n.strip().lower()==s)
                ws = wb[real]
                rows = list(ws.iter_rows(values_only=True))
                if not rows: return [], "jobs sheet empty"
                hdr = [str(h or "").strip() for h in rows[0]]
                idx = {h.strip().lower(): i for i,h in enumerate(hdr)}
                if all(n in idx for n in need):
                    J,A,D = (idx[n] for n in need)
                    out=[]
                    for r in rows[1:]:
                        if r is None: continue
                        jj = str((r[J] or "")).strip()
                        aa = str((r[A] or "")).strip()
                        dd = str((r[D] or "")).strip()
                        if jj or aa or dd:
                            out.append(f"{jj} - {aa} - {dd}")
                    return out, None
    except Exception as e:
        err = f"[jobs openpyxl] {e}"
    return [], err if 'err' in locals() else "jobs not found"

def __try_read_costs(xlsx):
    sheets = ["cost codes", "cost code list", "codes", "costcodes"]
    need   = ["cost code", "description"]
    # 1) pandas raw
    try:
        raw = xlsx.read_bytes()
        xl  = __pd.ExcelFile(raw)
        for s in sheets:
            if s in [n.lower() for n in xl.sheet_names]:
                real = next(n for n in xl.sheet_names if n.lower()==s)
                df = xl.parse(real)
                cols = __normalize_cols(df)
                if all(n in cols for n in need):
                    C, D = (cols[n] for n in need)
                    tmp = df[[C, D]].astype(str).fillna("")
                    return (tmp[C] + " - " + tmp[D]).tolist(), None
    except Exception as e:
        err = f"[cost raw] {e}"
    # 2) pandas normal
    try:
        xl, names = __excel_sheets(xlsx)
        if xl:
            for s in sheets:
                if s in [n.lower() for n in names]:
                    real = next(n for n in names if n.lower()==s)
                    df = xl.parse(real)
                    cols = __normalize_cols(df)
                    if all(n in cols for n in need):
                        C, D = (cols[n] for n in need)
                        tmp = df[[C, D]].astype(str).fillna("")
                        return (tmp[C] + " - " + tmp[D]).tolist(), None
    except Exception as e:
        err = f"[cost pandas] {e}"
    # 3) openpyxl
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filename=str(xlsx), read_only=True, data_only=True)
        names = [ws.title for ws in wb.worksheets]
        for s in sheets:
            if s in [n.strip().lower() for n in names]:
                real = next(n for n in names if n.strip().lower()==s)
                ws = wb[real]
                rows = list(ws.iter_rows(values_only=True))
                if not rows: return [], "cost sheet empty"
                hdr = [str(h or "").strip() for h in rows[0]]
                idx = {h.strip().lower(): i for i,h in enumerate(hdr)}
                if all(n in idx for n in need):
                    C,D = (idx[n] for n in need)
                    out=[]
                    for r in rows[1:]:
                        if r is None: continue
                        cc = str((r[C] or "")).strip()
                        dd = str((r[D] or "")).strip()
                        if cc or dd:
                            out.append(f"{cc} - {dd}")
                    return out, None
    except Exception as e:
        err = f"[cost openpyxl] {e}"
    return [], err if 'err' in locals() else "cost codes not found"

try:
    __root = __P(__file__).resolve().parents[1]   # project root
    __xlsx = __root / "Timesheet Apps.xlsx"
    if not job_options:
        job_options, __jerr = __try_read_jobs(__xlsx)
        if __jerr:
            __st.warning(f"Jobs list not loaded: {__jerr}")
    if not cost_options:
        cost_options, __cerr = __try_read_costs(__xlsx)
        if __cerr:
            __st.warning(f"Cost codes not loaded: {__cerr}")
    __st.caption(f"Loaded {len(job_options)} jobs, {len(cost_options)} cost codes.")
except Exception as __e:
    __st.warning(f"Job/Cost loader error: {__e}")
# ===== End injected loader =====
# ---- Fallback options loader (added) ----
# If job_options / cost_options were not created earlier, try to build them
# from "Timesheet Apps.xlsx" using common sheet/column names.
try:
    import pandas as _pd
    from pathlib import Path as _Path

    def _load_xlsx_df(_xlsx, candidates, needed_cols):
        try:
            xl = _pd.ExcelFile(_xlsx)
        except Exception:
            return None
        for nm in candidates:
            if nm in xl.sheet_names:
                try:
                    df = xl.parse(nm)
                    # normalize column names for matching
                    cols_norm = {c.strip(): c for c in df.columns if isinstance(c, str)}
                    # do we have a superset of the columns we need?
                    if all(any(k.lower() == want.lower() for k in cols_norm) for want in needed_cols):
                        return df
                except Exception:
                    continue
        return None

    # workbook next to the project root file locations
    _root = _Path(__file__).resolve().parents[1]
    _xlsx = _root / "Timesheet Apps.xlsx"

    # Build jobs if missing or empty
    if "job_options" not in locals() or not job_options:
        _jobs_df = _load_xlsx_df(
            _xlsx,
            ["Jobs", "Job List", "Jobs List", "JobsList"],
            ["Job Number", "Area Number", "Description"],
        )
        job_options = []
        if _jobs_df is not None and not _jobs_df.empty:
            # Be tolerant to header variants by matching case-insensitively
            def _col(df, name):
                for c in df.columns:
                    if isinstance(c, str) and c.strip().lower() == name.lower():
                        return c
                return name  # best-effort
            jn = _col(_jobs_df, "Job Number")
            an = _col(_jobs_df, "Area Number")
            ds = _col(_jobs_df, "Description")
            # stringify and build display text
            _tmp = _jobs_df[[jn, an, ds]].astype(str).fillna("")
            job_options = (_tmp[jn] + " - " + _tmp[an] + " - " + _tmp[ds]).tolist()

    # Build cost codes if missing or empty
    if "cost_options" not in locals() or not cost_options:
        _cost_df = _load_xlsx_df(
            _xlsx,
            ["Cost Codes", "Cost Code List", "Codes", "CostCodes"],
            ["Cost code", "Description"],
        )
        cost_options = []
        if _cost_df is not None and not _cost_df.empty:
            def _col(df, name):
                for c in df.columns:
                    if isinstance(c, str) and c.strip().lower() == name.lower():
                        return c
                return name
            cc = _col(_cost_df, "Cost code")
            ds = _col(_cost_df, "Description")
            _tmp = _cost_df[[cc, ds]].astype(str).fillna("")
            cost_options = (_tmp[cc] + " - " + _tmp[ds]).tolist()
except Exception:
    # Donâ€™t crash the page if workbook/sheets arenâ€™t available
    pass
# ---- End fallback options loader ----

import pandas as pd
from datetime import date

# Gate: require login
user = st.session_state.get("user_email")
if not user:
    st.warning("Please sign in on the Home page first.")
    if hasattr(st, "page_link"):
        st.page_link("streamlit_app.py", label="ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¾Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã¢â‚¬Â¦Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã¢â‚¬Â¦Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã¢â‚¬Â¦Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â Go to Home")
    st.stop()
st.sidebar.info(f"Signed in as: {user}")

from app.config import APP_NAME, EXCEL_FILENAME
from app.data.employee_loader import load_employee_list
from app.data.employee_details import load_employee_details
from app.data.reference_loader import load_job_options, load_costcode_options
from app.features.export_daily_time import export_daily_time

st.set_page_config(page_title=f"{APP_NAME}", layout="wide")
st.title("Timesheet Entry")

st.caption("Heartbeat: UI rendered. If this disappears after selecting an employee, tell me.")

# Date
col_date, _ = st.columns([1, 3])
with col_date:
    chosen_date = st.date_input("Date", value=date.today())

st.markdown("---")

# Job & Cost Code selectors
st.subheader("Job & Cost Code")

@st.cache_data(show_spinner=False)
def _jobs_cached():
    try:
        return load_job_options()
    except Exception:
        return []

@st.cache_data(show_spinner=False)
def _codes_cached():
    try:
        return load_costcode_options()
    except Exception:
        return []

jobs = _jobs_cached()
codes = _codes_cached()

c1, c2 = st.columns(2)
with c1:
    job_choice = job_options = (locals().get("job_display")
               or locals().get("job_options")
               or locals().get("jobs_display")
               or [])
job_options = (locals().get("job_display") or locals().get("job_options") or locals().get("jobs_display") or [])
job_choice  = st.selectbox("Job Number - Area Number - Description", job_options, key="job_choice")