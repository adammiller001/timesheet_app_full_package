import streamlit as st
import pandas as pd
from pathlib import Path
import io
import zipfile
import os
import re
import tempfile
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import openpyxl.styles
from shutil import copyfile
from time import sleep
import time
from app.style_utils import apply_watermark
from datetime import datetime, date
from typing import Optional

apply_watermark()

# Try to use your helpers; fall back gracefully if not present
try:
    from utils_jobs import (
        load_jobs_active,
        build_job_options,
        load_cost_options,
    )
    HAVE_UTILS = True
except Exception:
    HAVE_UTILS = False

# Import Google Sheets integration
try:
    from app.integrations.google_sheets import read_timesheet_data, get_sheets_manager
    HAVE_GOOGLE_SHEETS = True
except Exception:
    HAVE_GOOGLE_SHEETS = False

# Gate: require login
if not st.session_state.get("authenticated", False):
    st.warning("Please sign in on the Home page first.")
    st.stop()

user = st.session_state.get("user_email")
user_type = st.session_state.get("user_type", "User")

sheet_id_value = str(st.secrets.get("google_sheets_id", "")).strip() if 'google_sheets_id' in st.secrets else ""
service_account_email = ""
try:
    gs_section = dict(st.secrets.get("google_sheets", {}))
    service_account_email = str(gs_section.get("client_email", "")).strip()
except Exception:
    service_account_email = ""
if not st.session_state.get("_sheets_debug_shown", False):
    if not sheet_id_value or not service_account_email or not HAVE_GOOGLE_SHEETS:
        debug_id = f"{sheet_id_value[:6]}..." if sheet_id_value and len(sheet_id_value) > 10 else sheet_id_value
        debug_email = service_account_email if service_account_email else "missing"
        st.caption(f"Sheets config issue -> ID set: {bool(sheet_id_value)}, ID: {debug_id or 'n/a'}, service account: {debug_email}, gspread import: {HAVE_GOOGLE_SHEETS}")
    st.session_state['_sheets_debug_shown'] = True
GOOGLE_CONFIGURED = (
    "google_sheets_id" in st.secrets
    and str(st.secrets["google_sheets_id"]).strip()
)

if GOOGLE_CONFIGURED and not HAVE_GOOGLE_SHEETS:
    if not st.session_state.get("_google_dependency_warning_shown", False):
        st.warning("Google Sheets integration is configured, but required packages are missing. Install gspread and google-auth (e.g. run pip install -r requirements.txt) before exporting.")
        st.session_state["_google_dependency_warning_shown"] = True


# Initialize automatic data refresh trigger for truly dynamic dropdowns
if "auto_fresh_data" not in st.session_state:
    st.session_state.auto_fresh_data = True

XLSX = None

# Optional local Excel fallback path (set secrets.local_workbook_path to enable)
try:
    _local_workbook = str(st.secrets.get('local_workbook_path', '')).strip() if 'local_workbook_path' in st.secrets else ''
    if _local_workbook:
        candidate = Path(_local_workbook)
        if not candidate.is_absolute():
            candidate = Path.cwd() / candidate
        if candidate.exists():
            XLSX = candidate
except Exception:
    XLSX = None

TIME_DATA_COLUMNS = [
    "Job Number", "Job Area", "Date", "Name", "Trade Class",
    "Employee Number", "RT Hours", "OT Hours", "Description of work",
    "Comments", "Night Shift", "Premium Rate", "Subsistence Rate",
    "Travel Rate", "Indirect", "Cost Code"
]

# File modification time monitoring for automatic reloads
def check_file_modified():
    """Placeholder: Google Sheets updates are handled via cache refresh."""
    return False

# Check for file modifications
check_file_modified()

# Completely disable pandas caching
pd.set_option('io.hdf.default_format','table')
if hasattr(pd.io.common, '_maybe_convert_usecols'):
    pd.io.common._maybe_convert_usecols = lambda x: x  # Disable column caching

def safe_read_excel_force_fresh(file_path, sheet_name):
    """Legacy helper retained for compatibility; reads from Google Sheets."""
    return safe_read_excel(file_path, sheet_name, force_refresh=True)

def safe_read_excel(file_path, sheet_name, force_refresh=False):
    """Read worksheet data either from Google Sheets or a local Excel workbook."""
    sheet_id = str(st.secrets.get('google_sheets_id', '')).strip() if 'google_sheets_id' in st.secrets else ''
    if HAVE_GOOGLE_SHEETS and sheet_id:
        try:
            df = read_timesheet_data(sheet_name, force_refresh=force_refresh)
            if isinstance(df, pd.DataFrame) and not df.empty:
                df = df.copy()
                df.columns = [str(c).strip() for c in df.columns]
                return df
        except Exception:
            pass

    # Fallback to a local workbook if one was configured
    candidate_path = None
    try:
        if file_path:
            candidate_path = Path(file_path)
        elif XLSX:
            candidate_path = Path(XLSX)
    except Exception:
        candidate_path = None

    if candidate_path and candidate_path.exists():
        try:
            df = pd.read_excel(candidate_path, sheet_name=sheet_name)
            if isinstance(df, pd.DataFrame):
                df = df.copy()
                df.columns = [str(c).strip() for c in df.columns]
                return df
        except Exception:
            pass
    return pd.DataFrame()

@st.cache_data(show_spinner=False, ttl=60)
def _cached_sheet_data(sheet_name: str, cache_token: int, force_refresh: bool):
    try:
        df = read_timesheet_data(sheet_name, force_refresh=force_refresh)
        if isinstance(df, pd.DataFrame):
            return df
    except Exception as e:
        st.warning(f"Google Sheets read issue for {sheet_name}: {e}")
    return pd.DataFrame()


def _resolve_user_type(email: str, fallback: str) -> str:
    """Lookup user status from Users sheet; fall back gracefully."""
    if not email:
        return fallback
    try:
        users_df = smart_read_data("Users", force_refresh=True)
        if isinstance(users_df, pd.DataFrame) and not users_df.empty:
            users_df = users_df.copy()
            users_df.columns = [str(c).strip() for c in users_df.columns]
            email_col = None
            for cand in ("Email", "User Email", "Email Address", "E-mail", "Login", "User"):
                if cand in users_df.columns:
                    email_col = cand
                    break
            status_col = None
            for cand in ("Status", "Role", "Access Level", "User Type", "Type", "Permission"):
                if cand in users_df.columns:
                    status_col = cand
                    break
            if email_col is None and len(users_df.columns) > 0:
                email_col = users_df.columns[0]
            if status_col is None:
                if len(users_df.columns) >= 4:
                    status_col = users_df.columns[3]
                elif len(users_df.columns) > 0:
                    status_col = users_df.columns[-1]
            if email_col and status_col:
                normalized_email = str(email).strip().lower()
                matches = users_df[users_df[email_col].astype(str).str.strip().str.lower() == normalized_email]
                if not matches.empty:
                    value = str(matches.iloc[0][status_col]).strip()
                    if value:
                        return value
    except Exception:
        pass
    return fallback

user_type = _resolve_user_type(user, user_type)
st.session_state["user_type"] = user_type
st.sidebar.info(f"Signed in as: {user} ({user_type})")

if st.sidebar.button("Refresh All Dropdowns", use_container_width=True):
    st.session_state['force_fresh_data'] = True
    st.session_state['sheet_cache_token'] = st.session_state.get('sheet_cache_token', 0) + 1
    st.session_state['data_refresh_timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    st.rerun()


def smart_read_data(sheet_name, force_refresh=False):
    """Smart data reader that minimizes Google Sheets requests and falls back gracefully"""
    try:
        cache_token = st.session_state.get("sheet_cache_token", 0)
        if force_refresh:
            cache_token += 1
            st.session_state["sheet_cache_token"] = cache_token

        df = _cached_sheet_data(sheet_name, cache_token, force_refresh)
        if isinstance(df, pd.DataFrame) and not df.empty:
            return df.copy()

        # Fallback to Excel file if Google Sheets unavailable or empty
        return safe_read_excel(XLSX, sheet_name, force_refresh=True)
    except Exception as e:
        st.error(f"Failed to load {sheet_name}: {e}")
        return pd.DataFrame()


def _fetch_sheet_dataframe(primary_sheet: str, alt_names: Optional[tuple[str, ...]] = None, *, force_refresh: bool = True) -> pd.DataFrame:
    """Fetch a worksheet as a DataFrame with direct Google Sheets access and cached fallback."""
    candidates = [primary_sheet]
    if alt_names:
        candidates.extend(list(alt_names))
    sheet_id = str(st.secrets.get('google_sheets_id', '')).strip() if 'google_sheets_id' in st.secrets else ''
    if HAVE_GOOGLE_SHEETS and sheet_id:
        try:
            manager = get_sheets_manager()  # type: ignore[name-defined]
            worksheet, actual_title = manager.find_worksheet(candidates, sheet_id)
            target_title = actual_title or primary_sheet
            if worksheet or actual_title:
                df = manager.read_worksheet(target_title, sheet_id, force_refresh=force_refresh)
                if isinstance(df, pd.DataFrame) and not df.empty:
                    df = df.copy()
                    df.columns = [str(c).strip() for c in df.columns]
                    return df
        except Exception as exc:
            st.warning(f"Google Sheets read error for {primary_sheet}: {exc}")
    df = smart_read_data(primary_sheet, force_refresh=force_refresh)
    if isinstance(df, pd.DataFrame) and not df.empty:
        df = df.copy()
        df.columns = [str(c).strip() for c in df.columns]
        return df
    return pd.DataFrame()

def get_available_worksheets(_: object = None):
    """Get list of available worksheets in the configured Google Sheet"""
    sheet_id = st.secrets.get("google_sheets_id", "")
    if not sheet_id or not HAVE_GOOGLE_SHEETS:
        return []
    manager = get_sheets_manager()
    titles = []
    try:
        worksheets = manager._list_worksheets_http(sheet_id) if hasattr(manager, "_list_worksheets_http") else []
        if hasattr(manager, "spreadsheet") and manager.spreadsheet:
            worksheets = manager.spreadsheet.worksheets()
    except Exception:
        worksheets = []
    for ws in worksheets:
        if isinstance(ws, str):
            titles.append(ws)
        elif isinstance(ws, dict) and 'title' in ws:
            titles.append(ws['title'])
        else:
            try:
                titles.append(ws.title)
            except Exception:
                continue
    return titles
def _prepare_time_data_dataframe(df: Optional[pd.DataFrame]) -> pd.DataFrame:
    if df is None or df.empty:
        df = pd.DataFrame(columns=TIME_DATA_COLUMNS)
    else:
        df = df.copy()
    for col in TIME_DATA_COLUMNS:
        if col not in df.columns:
            df[col] = ''
    ordered_cols = TIME_DATA_COLUMNS + [col for col in df.columns if col not in TIME_DATA_COLUMNS]
    df = df[ordered_cols]
    if 'Date' in df.columns and not df.empty:
        try:
            df['Date'] = pd.to_datetime(df['Date']).dt.strftime('%Y-%m-%d')
        except Exception:
            df['Date'] = df['Date'].astype(str)
    df = df.applymap(_normalize_sheet_value)
    return df

def _load_time_data_from_excel() -> pd.DataFrame:
    df = safe_read_excel(None, 'Time Data', force_refresh=True)
    if isinstance(df, pd.DataFrame):
        df.columns = [str(col).strip() for col in df.columns]
        return _prepare_time_data_dataframe(df)
    return pd.DataFrame(columns=TIME_DATA_COLUMNS)

def _write_time_data_to_excel(df: pd.DataFrame) -> bool:
    """Legacy shim: local workbook writes are no longer used."""
    return True

def _append_time_data_to_excel(new_data_df: pd.DataFrame) -> bool:
    """Legacy shim to preserve previous workflow."""
    return True

def _load_latest_time_data_for_sync() -> tuple[pd.DataFrame, str]:
    try:
        if GOOGLE_CONFIGURED and HAVE_GOOGLE_SHEETS:
            st.session_state['sheet_cache_token'] = st.session_state.get('sheet_cache_token', 0) + 1
            latest = smart_read_data("Time Data", force_refresh=True)
            if isinstance(latest, pd.DataFrame):
                prepared = _prepare_time_data_dataframe(latest)
                prepared = _enrich_with_employee_details(prepared)
                if not prepared.empty or len(prepared.columns) > 0:
                    return prepared, "google"
                return prepared, "google"
    except Exception:
        pass

    fallback = _load_time_data_from_excel()
    fallback = _prepare_time_data_dataframe(fallback)
    fallback = _enrich_with_employee_details(fallback)
    return fallback, "excel"

def get_time_data_from_session(_date_filter=None):
    """Get time data from session state with optional date filtering"""
    try:
        data = st.session_state.session_time_data.copy()

        if _date_filter and "Date" in data.columns and not data.empty:
            data["Date"] = pd.to_datetime(data["Date"])
            data = data[data["Date"].dt.strftime("%Y-%m-%d") == _date_filter]

        return data
    except Exception as e:
        st.error(f"Error reading session time data: {e}")
        return pd.DataFrame()

def _normalize_sheet_value(value):
    if value is None:
        return ''
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, bool):
        return 'TRUE' if value else 'FALSE'
    if isinstance(value, (int, float)):
        if pd.isna(value):
            return ''
        return ("{0:.15g}".format(float(value))).rstrip('.0') if float(value).is_integer() else "{0:.15g}".format(float(value))
    if isinstance(value, pd.Timestamp):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')
    try:
        if pd.isna(value):
            return ''
    except Exception:
        pass
    return str(value).strip()


def _build_sheet_row(row_dict, headers):
    normalized = []
    for header in headers:
        value = row_dict.get(header)
        normalized.append(_normalize_sheet_value(value))
    return normalized


def _sync_time_data_to_google(new_data_df: pd.DataFrame) -> bool:
    """Push new time data rows to Google Sheets when available"""
    if new_data_df.empty:
        return True

    if not (
        HAVE_GOOGLE_SHEETS
        and "google_sheets_id" in st.secrets
        and st.secrets["google_sheets_id"]
    ):
        return True

    try:
        sheet_id = st.secrets["google_sheets_id"]
        manager = get_sheets_manager()
        sheet_candidates = ("Time Data", "TimeData")
        worksheet, actual_title = manager.find_worksheet(sheet_candidates, sheet_id)

        if not worksheet or not actual_title:
            st.warning("Time Data worksheet not found in Google Sheets. Please ensure a tab named 'Time Data' exists.")
            return False

        df_to_sync = new_data_df.copy()
        if df_to_sync.empty:
            return True

        df_to_sync.columns = [str(col).strip() for col in df_to_sync.columns]
        df_to_sync = df_to_sync[[col for col in df_to_sync.columns if col]]
        df_to_sync = df_to_sync.where(pd.notnull(df_to_sync), None)

        header_values = worksheet.row_values(1)
        headers = [str(cell).strip() for cell in header_values if str(cell).strip()]

        if not headers:
            all_values = worksheet.get_all_values()
            for row in all_values:
                cleaned = [str(cell).strip() for cell in row if str(cell).strip()]
                if cleaned:
                    headers = [str(cell).strip() for cell in row if str(cell).strip()]
                    break

        if not headers:
            manager.write_worksheet(actual_title, df_to_sync, sheet_id)
            return True

        def _norm(col_name: str) -> str:
            return ''.join(ch for ch in str(col_name).strip().lower() if ch.isalnum())

        df_lookup = {_norm(col): col for col in df_to_sync.columns}
        rows_to_append = []

        for _, row in df_to_sync.iterrows():
            row_values = []
            for header in headers:
                source_col = df_lookup.get(_norm(header))
                value = row[source_col] if source_col else None
                if pd.isna(value):
                    value = None
                elif isinstance(value, pd.Timestamp):
                    value = value.strftime("%Y-%m-%d")
                elif hasattr(value, "item") and not isinstance(value, (str, bytes)):
                    try:
                        value = value.item()
                    except Exception:
                        value = str(value)
                row_values.append(value)
            rows_to_append.append(row_values)

        if not rows_to_append:
            return True

        return manager.append_rows(actual_title, rows_to_append, sheet_id)

    except Exception as e:
        st.error(f"Failed to sync Time Data to Google Sheets: {e}")
        return False


def _replace_time_data_in_google(updated_df: pd.DataFrame) -> bool:
    """Write the provided DataFrame to the Google Sheets Time Data worksheet"""
    if not (
        HAVE_GOOGLE_SHEETS
        and "google_sheets_id" in st.secrets
        and st.secrets["google_sheets_id"]
    ):
        return True

    try:
        sheet_id = st.secrets["google_sheets_id"]
        manager = get_sheets_manager()
        sheet_candidates = ("Time Data", "TimeData")
        worksheet, actual_title = manager.find_worksheet(sheet_candidates, sheet_id)

        if not worksheet or not actual_title:
            st.warning("Time Data worksheet not found in Google Sheets. Please ensure a tab named 'Time Data' exists.")
            return False

        header_values = worksheet.row_values(1)
        headers = [str(cell).strip() for cell in header_values if str(cell).strip()]

        df_to_write = _prepare_time_data_dataframe(updated_df)

        if headers:
            for header in headers:
                if header not in df_to_write.columns:
                    df_to_write[header] = ''
            df_to_write = df_to_write[headers]
        elif not df_to_write.empty:
            headers = list(df_to_write.columns)
        else:
            headers = TIME_DATA_COLUMNS.copy()
            for header in headers:
                if header not in df_to_write.columns:
                    df_to_write[header] = ''
            df_to_write = df_to_write[headers]

        success = manager.write_worksheet(actual_title, df_to_write, sheet_id)
        return bool(success)
    except Exception as e:
        st.error(f"Failed to update Time Data in Google Sheets: {e}")
        return False




def _is_blank_value(val) -> bool:
    if val is None:
        return True
    try:
        import math
        if isinstance(val, (float, int)) and math.isnan(val):
            return True
    except Exception:
        pass
    text = str(val).strip()
    return text == '' or text.lower() in {"nan", "none"}

def _enrich_with_employee_details(df: pd.DataFrame) -> pd.DataFrame:
    """Fill Night Shift and rate columns from Employee List when missing."""
    if df is None or df.empty:
        return df
    try:
        employee_df = smart_read_data("Employee List", force_refresh=True)
        if not isinstance(employee_df, pd.DataFrame) or employee_df.empty:
            return df
        employee_df = employee_df.copy()
        employee_df.columns = [str(c).strip() for c in employee_df.columns]
        info = {}
        name_col = _find_col(employee_df, ["Employee Name", "Name"]) or "Employee Name"
        night_col = _find_col(employee_df, ["Night Shift", "NightShift", "Nightshift", "Night"]) or "Night Shift"
        premium_col = _find_col(employee_df, ["Premium Rate", "Premium"]) or "Premium Rate"
        subsistence_col = _find_col(employee_df, ["Subsistence Rate", "Subsistence", "Column H"]) or "Subsistence Rate"
        travel_col = _find_col(employee_df, ["Travel Rate", "Travel"]) or "Travel Rate"
        for _, emp_row in employee_df.iterrows():
            name = str(emp_row.get(name_col, "")).strip()
            if not name:
                continue
            night_val = _normalize_night_flag(emp_row.get(night_col, ""))
            premium_val = str(emp_row.get(premium_col, "") or "").strip()
            subsistence_val = str(emp_row.get(subsistence_col, "") or "").strip()
            travel_val = str(emp_row.get(travel_col, "") or "").strip()
            info[name] = {
                "night": night_val,
                "premium": '' if _is_blank_value(premium_val) else premium_val,
                "subsistence": '' if _is_blank_value(subsistence_val) else subsistence_val,
                "travel": '' if _is_blank_value(travel_val) else travel_val
            }
        if not info:
            return df
        df = df.copy()
        for idx, row in df.iterrows():
            name = str(row.get("Name", "")).strip()
            if not name:
                continue
            details = info.get(name)
            if not details:
                continue
            df.at[idx, "Night Shift"] = details.get("night", "") or ""
            df.at[idx, "Premium Rate"] = details.get("premium", "")
            df.at[idx, "Subsistence Rate"] = details.get("subsistence", "")
            df.at[idx, "Travel Rate"] = details.get("travel", "")
        return df
    except Exception:
        return df


def save_to_session(new_rows):
    """Save new rows to session, then persist entire dataset to Excel/Google when configured"""
    try:
        if not new_rows:
            return False

        new_data_df = pd.DataFrame(new_rows)
        if new_data_df.empty:
            return False

        new_data_df = _enrich_with_employee_details(new_data_df)

        refreshed_df, source = _load_latest_time_data_for_sync()
        existing_df = st.session_state.get("session_time_data")

        if source == "google":
            combined_source = refreshed_df
        else:
            if isinstance(existing_df, pd.DataFrame) and not existing_df.empty:
                combined_source = existing_df
            else:
                combined_source = refreshed_df

        combined_source = _enrich_with_employee_details(combined_source)
        if combined_source is None or combined_source.empty:
            combined_source = pd.DataFrame(columns=TIME_DATA_COLUMNS.copy())

        updated_df = pd.concat([combined_source, new_data_df], ignore_index=True)
        updated_df = _enrich_with_employee_details(updated_df)
        updated_df = _prepare_time_data_dataframe(updated_df)
        st.session_state.session_time_data = updated_df

        excel_synced = _write_time_data_to_excel(updated_df)
        if not excel_synced:
            st.warning("Added lines locally, but could not update local Time Data worksheet.")

        google_synced = True
        if GOOGLE_CONFIGURED and HAVE_GOOGLE_SHEETS:
            google_synced = _sync_time_data_to_google(new_data_df)
            if not google_synced:
                google_synced = _replace_time_data_in_google(updated_df)
            if not google_synced:
                st.warning("Added lines locally, but could not update Google Sheets Time Data.")

        st.session_state['sheet_cache_token'] = st.session_state.get('sheet_cache_token', 0) + 1
        return excel_synced or google_synced
    except Exception as e:
        st.error(f"Error saving to Time Data: {e}")
        return False


st.markdown("### Timesheet Entry")



# Force fresh data reload if refresh was requested
if st.session_state.get("force_fresh_data", False):
    st.info(f"🔄 Loading fresh data... (Timestamp: {st.session_state.get('data_refresh_timestamp', 'N/A')})")
    # Reset flag after use
    st.session_state.force_fresh_data = False

# Status indicators removed - Google Sheets working properly

# Add live data verification - now optional since issues are resolved
if False:
    try:
        st.write("**Testing direct Excel read vs safe_read_excel:**")

        # Try direct pandas read
        try:
            direct_read = smart_read_data("Employee List", force_refresh=True)
            st.write(f"Google read (direct): {len(direct_read)} rows")
        except Exception as e:
            st.write(f"Direct read failed: {e}")

        # Try our safe read
        emp_debug = smart_read_data("Employee List", force_refresh=True)
        st.write(f"**safe_read_excel: {len(emp_debug)} rows**")
        st.write("**All columns:**", list(emp_debug.columns))

        # Check if there are any NaN values in Employee Name that might cause issues
        if "Employee Name" in emp_debug.columns:
            nan_count = emp_debug["Employee Name"].isna().sum()
            st.write(f"**NaN values in Employee Name column: {nan_count}**")

            # Show any rows with NaN names
            nan_rows = emp_debug[emp_debug["Employee Name"].isna()]
            if not nan_rows.empty:
                st.write("**Rows with NaN Employee Names:**")
                st.dataframe(nan_rows)

        # Show all employee names
        name_col = None
        for col in ["Employee Name", "Name", "Employee", "Full Name"]:
            if col in emp_debug.columns:
                name_col = col
                break

        if name_col:
            st.write(f"**All employees in {name_col} column:**")
            st.write(f"**Expected: 11 employees, Actually loaded: {len(emp_debug)}**")

            # Show all rows with their original Excel row numbers if possible
            for idx, row in emp_debug.iterrows():
                name = str(row[name_col])
                person_num = row.get('Person Number', 'N/A')
                active_val = row.get('Active', 'N/A')
                st.write(f"Row {idx+2}: '{name}' (Person: {person_num}, Active: {active_val})")

            # Check specifically for Graham with various spellings
            graham_variations = ['GRAHAM', 'GRAEME', 'GRAHM', 'ST HILAIRE', 'HILAIRE']
            for variation in graham_variations:
                graham_check = emp_debug[emp_debug[name_col].astype(str).str.contains(variation, case=False, na=False)]
                if not graham_check.empty:
                    st.success(f"✅ Found '{variation}' in: {graham_check[name_col].iloc[0]}")
                else:
                    st.write(f"⚪ No match for '{variation}'")

            # Show the full dataframe for inspection
            st.write("**Full Employee DataFrame:**")
            st.dataframe(emp_debug)

            # Check specifically for Graham
            graham_check = emp_debug[emp_debug[name_col].astype(str).str.contains("GRAHAM", case=False, na=False)]
            if not graham_check.empty:
                st.success(f"✅ GRAHAM FOUND in Excel: {graham_check[name_col].iloc[0]}")
            else:
                st.error("❌ GRAHAM NOT FOUND in Excel Employee List")

        # Check Active column
        if "Active" in emp_debug.columns:
            st.write("**Active column data types and values:**")
            st.write(f"Column data type: {emp_debug['Active'].dtype}")
            st.write("Unique values and their types:")
            for val in emp_debug['Active'].unique():
                st.write(f"  '{val}' (type: {type(val).__name__})")

            # Show filtering results
            bool_filter = (emp_debug["Active"] == True)
            str_filter = emp_debug["Active"].astype(str).str.upper().isin(["TRUE", "YES", "Y", "1"])
            combined_filter = bool_filter | str_filter

            st.write(f"Boolean filter (== True): {bool_filter.sum()} employees")
            st.write(f"String filter (TRUE/YES/Y/1): {str_filter.sum()} employees")
            st.write(f"Combined filter result: {combined_filter.sum()} employees")

            filtered_df = emp_debug[combined_filter]
            if name_col:
                st.write("**Employees after Active filtering:**")
                for name in filtered_df[name_col].astype(str).tolist():
                    st.write(f"  - '{name}'")
        else:
            st.warning("No 'Active' column found")

    except Exception as e:
        st.error(f"Employee debug error: {e}")
        import traceback
        st.code(traceback.format_exc())

    st.write("## **DEBUGGING COST CODES**")
    try:
        cost_debug = smart_read_data("Cost Codes", force_refresh=True)
        st.write(f"**RAW DATA LOADED - Total rows: {len(cost_debug)}**")
        st.write("**All columns:**", list(cost_debug.columns))

        if "Active" in cost_debug.columns:
            st.write("**Active column analysis:**")
            st.write(f"Data type: {cost_debug['Active'].dtype}")
            active_vals = cost_debug['Active'].value_counts()
            st.write("Value counts:", dict(active_vals))

            # Show which items are FALSE
            false_items = cost_debug[cost_debug['Active'] == False]
            if not false_items.empty:
                st.write("**Items marked as FALSE:**")
                for idx, row in false_items.iterrows():
                    code_val = row.get('Cost Code', row.get('Code', 'Unknown'))
                    desc_val = row.get('Description', row.get('DESC', 'No description'))
                    st.write(f"  - {code_val} - {desc_val}")

            # Test filtering
            bool_filter = (cost_debug["Active"] == True)
            str_filter = cost_debug["Active"].astype(str).str.upper().isin(["TRUE", "YES", "Y", "1"])
            combined_filter = bool_filter | str_filter

            st.write(f"Items that should be filtered OUT (Active=False): {len(cost_debug) - combined_filter.sum()}")
            st.write(f"Items that should show in dropdown: {combined_filter.sum()}")

            # Specific debug for the problematic cost code
            st.write("**SPECIFIC DEBUG: 00002-170-53**")
            specific_code = cost_debug[cost_debug['Cost Code'].astype(str).str.contains('00002-170-53', na=False)]
            if not specific_code.empty:
                row = specific_code.iloc[0]
                st.write(f"Found: {row['Cost Code']} - {row['Description']}")
                st.write(f"Active value: {row['Active']} (type: {type(row['Active']).__name__})")
                st.write(f"Passes boolean filter: {row['Active'] == True}")
                st.write(f"Passes string filter: {str(row['Active']).upper() in ['TRUE', 'YES', 'Y', '1']}")
            else:
                st.write("❌ 00002-170-53 NOT FOUND in cost codes data")

    except Exception as e:
        st.error(f"Cost codes debug error: {e}")
        import traceback
        st.code(traceback.format_exc())

# --- Date ---
date_val = st.date_input("Date", value=pd.Timestamp.today().date(), format="YYYY/MM/DD", key="date_val")

# Form reset mechanism
if "form_counter" not in st.session_state:
    st.session_state.form_counter = 0

# Initialize session-based Time Data storage
if "session_time_data" not in st.session_state:
    # Load initial data with Google Sheets support
    try:
        initial_data = smart_read_data("Time Data")
        initial_data = _enrich_with_employee_details(initial_data)
    except Exception as e:
        st.warning(f"Failed to load Time Data from configured sources: {e}")
        initial_data = pd.DataFrame()

    if isinstance(initial_data, pd.DataFrame) and (not initial_data.empty or len(initial_data.columns) > 0):
        st.session_state.session_time_data = initial_data
    else:
        st.session_state.session_time_data = pd.DataFrame(columns=TIME_DATA_COLUMNS.copy())

# --- Helper functions ---
def _find_col(df: pd.DataFrame, candidates):
    for c in candidates:
        if c in df.columns:
            return c
    return None

def _pad_area(val: object) -> str:
    """Ensure area is 3 digits with zero padding"""
    s = str(val).strip()
    if not s or s.lower() == "nan":
        return "000"
    m = re.search(r"\d+", s)
    if m:
        d = m.group(0)
        return d.zfill(3)
    return s

def _is_truthy(value) -> bool:
    """Return True if the value represents an affirmative flag."""
    str_val = str(value).strip().upper()
    if isinstance(value, bool):
        return value
    if str_val in {"TRUE", "YES", "Y", "1", "ON"}:
        return True
    try:
        return float(str_val) == 1.0
    except Exception:
        return False



def _build_job_options_local(df: pd.DataFrame):
    """Build job options with robust column matching and padding."""
    if df is None or df.empty:
        return []
    job_c = _find_col(df, [
        "Job Number", "JOB #", "Job #", "Job", "JobNumber", "Job No", "Job_No"
    ])
    area_c = _find_col(df, [
        "Area Number", "AREA #", "Area #", "AREA#", "Area", "Area No"
    ])
    desc_c = _find_col(df, [
        "Description", "DESCRIPTION", "PROJECT NAME", "Project Name", "Job Description", "Description of Work"
    ])
    if not job_c and len(df.columns) > 0:
        job_c = df.columns[0]
    if not area_c and len(df.columns) > 1:
        area_c = df.columns[1]
    if not desc_c and len(df.columns) > 2:
        desc_c = df.columns[2]
    if not job_c:
        return []
    out = []
    for _, row in df.iterrows():
        j = str(row.get(job_c, "") or "").strip()
        a = _pad_area(row.get(area_c, "")) if area_c else "000"
        d = str(row.get(desc_c, "") or "").strip() if desc_c else ""
        if j or a or d:
            label = f"{j} - {a} - {d}" if d else f"{j} - {a}"
            out.append(label.strip(" -"))
    return sorted(pd.Series(out).dropna().astype(str).unique().tolist())


def _is_quarter_hour(value: float) -> bool:
    return abs((value * 4) - round(value * 4)) < 1e-6


def _parse_hours_input(raw_value: str) -> Optional[float]:
    if raw_value is None:
        return 0.0
    raw_value = str(raw_value).strip()
    if raw_value == "":
        return 0.0
    try:
        value = float(raw_value)
    except ValueError:
        return None
    if value < 0:
        return None
    return round(value, 2)






def _load_active_job_options() -> tuple[list[str], int, int]:
    """Load job dropdown options and report total/active row counts."""
    total_rows = 0
    active_rows = 0
    try:
        jobs_df = _fetch_sheet_dataframe("Job Numbers", ("Jobs",), force_refresh=True)
        if isinstance(jobs_df, pd.DataFrame) and not jobs_df.empty:
            jobs_df = jobs_df.copy()
            total_rows = len(jobs_df)
            cols = list(jobs_df.columns)
            job_col = cols[2] if len(cols) > 2 else cols[0]
            area_col = cols[3] if len(cols) > 3 else cols[min(1, len(cols)-1)]
            desc_col = cols[5] if len(cols) > 5 else cols[min(2, len(cols)-1)]
            active_col = cols[6] if len(cols) > 6 else None
            if active_col and active_col in jobs_df.columns:
                mask = jobs_df[active_col].apply(_is_truthy)
                active_rows = int(mask.sum())
                if active_rows:
                    jobs_df = jobs_df[mask]
            options = []
            for _, row in jobs_df.iterrows():
                job = str(row.get(job_col, "") or "").strip()
                area = _pad_area(row.get(area_col, "")) if area_col else "000"
                desc = str(row.get(desc_col, "") or "").strip() if desc_col else ""
                if job or area or desc:
                    label = f"{job} - {area} - {desc}" if desc else f"{job} - {area}"
                    options.append(label.strip(' -'))
            if options:
                options = sorted(pd.Series(options).dropna().astype(str).unique().tolist())
                return options, total_rows, active_rows
    except Exception as exc:
        st.warning(f"Could not load Job Numbers sheet: {exc}")
    return [], total_rows, active_rows

job_options, job_total_rows, job_active_rows = _load_active_job_options()
if not job_options:
    if job_total_rows == 0:
        st.warning("Job Numbers sheet returned 0 rows. Confirm the worksheet has data and the service account can access it.")
    else:
        st.warning(f"No active jobs found. Check the Job Numbers sheet and ensure new rows are marked active (column G). Rows fetched: {job_total_rows}, active flagged: {job_active_rows}.")

job_choice = st.selectbox(
    "Job Number - Area Number - Description",
    job_options,
    index=None,
    placeholder="Select a job...",
    key=f"job_choice_{st.session_state.form_counter}"
)









def _load_active_cost_codes() -> tuple[list[str], int, int]:
    """Return cost code options and counts of total/active rows."""
    total_rows = 0
    active_rows = 0
    try:
        cost_df = _fetch_sheet_dataframe("Cost Codes", ("CostCodes",), force_refresh=True)
        if isinstance(cost_df, pd.DataFrame) and not cost_df.empty:
            cost_df = cost_df.copy()
            total_rows = len(cost_df)
            cost_df.columns = [str(c).strip() for c in cost_df.columns]
            cols = list(cost_df.columns)
            code_col = cols[0]
            desc_col = cols[1] if len(cols) > 1 else cols[0]
            active_col = cols[2] if len(cols) > 2 else None
            if active_col and active_col in cost_df.columns:
                mask = cost_df[active_col].apply(_is_truthy)
                active_rows = int(mask.sum())
                if active_rows:
                    cost_df = cost_df[mask]
            descriptions = []
            for _, row in cost_df.iterrows():
                code = str(row.get(code_col, "") or "").strip()
                desc = str(row.get(desc_col, "") or "").strip() if desc_col else ""
                if code:
                    descriptions.append(f"{code} - {desc}" if desc else code)
            if descriptions:
                descriptions = sorted(pd.Series(descriptions).dropna().astype(str).unique().tolist())
                return descriptions, total_rows, active_rows
    except Exception as exc:
        st.warning(f"Could not load Cost Codes sheet: {exc}")
    return [], total_rows, active_rows

cost_options, cost_total_rows, cost_active_rows = _load_active_cost_codes()
if not cost_options:
    if cost_total_rows == 0:
        st.warning("Cost Codes sheet returned 0 rows. Confirm the worksheet has data and the service account can access it.")
    else:
        st.warning(f"No active cost codes found. Check the Cost Codes sheet and ensure new rows are marked active (column C). Rows fetched: {cost_total_rows}, active flagged: {cost_active_rows}.")

cost_choice = st.selectbox(
    "Cost Code - Description",
    cost_options,
    index=None,
    placeholder="Select a cost code...",
    key=f"cost_choice_{st.session_state.form_counter}"
)


# --- Employees (simple multiselect dropdown only) ---
employee_total_rows = 0
employee_active_rows = 0
try:
    _emp_df_raw = _fetch_sheet_dataframe("Employee List", ("Employees",), force_refresh=True)
    if isinstance(_emp_df_raw, pd.DataFrame) and not _emp_df_raw.empty:
        _emp_df = _emp_df_raw.copy()
        employee_total_rows = len(_emp_df)
        _emp_df.columns = [str(c).strip() for c in _emp_df.columns]
        cols = list(_emp_df.columns)
        name_col = cols[2] if len(cols) > 2 else cols[0]
        active_col = cols[10] if len(cols) > 10 else None
        if active_col and active_col in _emp_df.columns:
            mask = _emp_df[active_col].apply(_is_truthy)
            employee_active_rows = int(mask.sum())
            if employee_active_rows:
                _emp_df = _emp_df[mask]
        EMP_NAME_COL = name_col
        _employee_options = sorted(_emp_df[EMP_NAME_COL].dropna().astype(str).unique().tolist())
    else:
        raise ValueError("Employee sheet empty")
except Exception:
    _emp_df = pd.DataFrame()
    _employee_options = []
    EMP_NAME_COL = "Employee Name"
    employee_total_rows = 0
    employee_active_rows = 0

if not _employee_options:
    if employee_total_rows == 0:
        st.warning("Employee List sheet returned 0 rows. Confirm the worksheet has data and the service account can access it.")
    else:
        st.warning(f"No active employees available. Rows fetched: {employee_total_rows}, active flagged: {employee_active_rows}.")

selected_employees = st.multiselect(
    "Employees",
    options=_employee_options,
    default=[],
    placeholder="Select one or more employees...",
    key=f"selected_employees_{st.session_state.form_counter}"
)




def _normalize_night_flag(raw_value) -> str:
    if raw_value is None:
        return ""
    value = str(raw_value).strip()
    if not value or value.lower() in {"nan", "none"}:
        return ""
    if value.upper() in {"Y", "YES", "TRUE", "1"}:
        return "Y"
    return ""

def night_flag_for(_name: str) -> str:
    """Return 'Y' if employee has night shift flag, otherwise empty string"""
    try:
        if _emp_df.empty or EMP_NAME_COL not in _emp_df.columns:
            return ""
        _row = _emp_df.loc[_emp_df[EMP_NAME_COL].astype(str) == str(_name)]
        if _row.empty:
            return ""
        return _normalize_night_flag(_row.iloc[0].get("Night Shift", ""))
    except Exception:
        return ""

# --- Hours Input ---
cols = st.columns(3)
with cols[0]:
    rt_hours_raw = st.text_input("RT Hours", placeholder="Enter RT hours", key=f"rt_hours_{st.session_state.form_counter}")
    rt_hours = _parse_hours_input(rt_hours_raw)
    rt_hours_valid = True
    if rt_hours is None:
        rt_hours_valid = False
        if str(rt_hours_raw).strip():
            st.error("Enter a non-negative number (increments of 0.25).")
        rt_hours_value = 0.0
    else:
        if not _is_quarter_hour(rt_hours):
            st.error("Hours must be entered in 0.25 increments.")
            rt_hours_valid = False
        rt_hours_value = rt_hours if rt_hours_valid else 0.0

with cols[1]:
    ot_hours_raw = st.text_input("OT Hours", placeholder="Enter OT hours", key=f"ot_hours_{st.session_state.form_counter}")
    ot_hours = _parse_hours_input(ot_hours_raw)
    ot_hours_valid = True
    if ot_hours is None:
        ot_hours_valid = False
        if str(ot_hours_raw).strip():
            st.error("Enter a non-negative number (increments of 0.25).")
        ot_hours_value = 0.0
    else:
        if not _is_quarter_hour(ot_hours):
            st.error("Hours must be entered in 0.25 increments.")
            ot_hours_valid = False
        ot_hours_value = ot_hours if ot_hours_valid else 0.0

with cols[2]:
    st.write("")

comments = st.text_input("Comments", value="", key=f"comments_{st.session_state.form_counter}")

def _parse_job(choice: str):
    """Parse job choice into job number, area, and description"""
    if not choice:
        return "", "", ""
    parts = choice.split(" - ", 2)
    parts += [""] * (3 - len(parts))
    return parts[0], parts[1], parts[2]

# --- Add line ---
hours_valid = rt_hours_valid and ot_hours_valid
positive_hours = (rt_hours_value > 0) or (ot_hours_value > 0)
add_disabled = not (job_choice and cost_choice and selected_employees and hours_valid and positive_hours)

if st.button("Add line", type="primary", disabled=add_disabled):
    with st.spinner("Adding entries..."):
        job_num, job_area, job_desc = _parse_job(job_choice)
        
        try:
            employee_lookup = {}
            if not _emp_df.empty and EMP_NAME_COL in _emp_df.columns:
                for _, row in _emp_df.iterrows():
                    name = str(row.get(EMP_NAME_COL, ""))
                    employee_lookup[name] = {
                        'emp_num': str(row.get("Person Number", "") or ""),
                        'trade': str(row.get("Override Trade Class", "") or "") or str(row.get("Trade Class", "") or ""),
                        'premium': str(row.get("Premium Rate", "") or ""),
                        'subsistence': str(row.get("Subsistence Rate", "") or ""),
                        'travel': str(row.get("Travel Rate", "") or ""),
                        'indirect': str(row.get("Indirect / Direct", "")).strip().upper() == "INDIRECT",
                        'night': _normalize_night_flag(row.get("Night Shift", ""))
                    }
            
            new_rows = []
            date_str = pd.to_datetime(date_val).strftime("%Y-%m-%d")
            cost_code = cost_choice.split(" - ", 1)[0] if cost_choice else ""
            
            for emp_name in selected_employees:
                emp_data = employee_lookup.get(emp_name, {
                    'emp_num': '', 'trade': '', 'premium': '', 'subsistence': '', 'travel': '', 'indirect': False, 'night': ''
                })

                new_row = {
                    "Job Number": job_num,
                    "Job Area": job_area,
                    "Date": date_str,
                    "Name": emp_name,
                    "Trade Class": emp_data['trade'],
                    "Employee Number": emp_data['emp_num'],
                    "RT Hours": rt_hours_value,
                    "OT Hours": ot_hours_value,
                    "Description of work": job_desc,
                    "Comments": comments,
                    "Night Shift": emp_data['night'],
                    "Premium Rate": emp_data['premium'],
                    "Subsistence Rate": emp_data['subsistence'],
                    "Travel Rate": emp_data['travel'],
                    "Indirect": emp_data['indirect'],
                    "Cost Code": cost_code,
                }
                new_rows.append(new_row)
            
            success = save_to_session(new_rows)

            if success:
                current_counter = st.session_state.form_counter
                for suffix in ("job_choice", "cost_choice", "selected_employees", "rt_hours", "ot_hours", "comments"):
                    key = f"{suffix}_{current_counter}"
                    if key in st.session_state:
                        del st.session_state[key]

                st.success(f"Added {len(new_rows)} line(s) to Time Data.")
                # Clear form by incrementing counter (changes all widget keys)
                st.session_state.form_counter = current_counter + 1
                st.rerun()
            else:
                st.error("Failed to save data. Please try again.")
                
        except Exception as e:
            st.error(f"Could not save to Time Data: {e}")

# --- Display current Time Data with proper date filtering ---
st.divider()
col1, col2 = st.columns([3, 1])
with col1:
    st.subheader("Current Time Data")
with col2:
    if st.button("🔄 Refresh Data", help="Clear cache and reload Time Data"):
        refreshed_df, _ = _load_latest_time_data_for_sync()
        if isinstance(refreshed_df, pd.DataFrame):
            st.session_state.session_time_data = refreshed_df
        st.session_state['sheet_cache_token'] = st.session_state.get('sheet_cache_token', 0) + 1
        st.success("Time Data reloaded from source.")
        st.rerun()

try:
    selected_date_str = pd.to_datetime(date_val).strftime("%Y-%m-%d")
    filtered_data = get_time_data_from_session(selected_date_str)

    total_data = get_time_data_from_session(None)
    total_entries = len(total_data) if not total_data.empty else 0
    filtered_entries = len(filtered_data) if not filtered_data.empty else 0

    if user_type.upper() != "ADMIN":
        st.info("Current Time Data is available to administrators only.")
        st.stop()

    # Show helpful message when no data exists
    if total_entries == 0:
        available_sheets = get_available_worksheets(XLSX)
        if available_sheets and 'Time Data' not in available_sheets:
            st.error(f"'Time Data' worksheet not found! Available worksheets: {', '.join(available_sheets)}")
        elif total_entries == 0:
            st.info("Time Data worksheet is empty. Add some entries using the form above.")

    # Show message when no data exists for selected date but keep filtering strict
    if total_entries > 0 and filtered_entries == 0:
        st.info(f"No entries found for {date_val}. Only entries matching this date will be shown.")
        # Keep filtered_data empty - don't show all entries from other dates

    st.caption(f"Showing {filtered_entries} of {total_entries} total entries for {date_val}")

    if not filtered_data.empty:
        st.dataframe(filtered_data, use_container_width=True, hide_index=True)

        # Only show Delete Entries section to Admin users
        if user_type.upper() == "ADMIN":
            st.subheader("Delete Entries")

            delete_tabs = st.tabs(["Delete Multiple", "Delete All"])

            with delete_tabs[0]:
                option_labels = {}
                option_keys = []
                for display_idx, (actual_index, row) in enumerate(filtered_data.iterrows(), start=1):
                    label = (
                        f"Row {display_idx}: {row['Name']} - {row['Date']} - "
                        f"{row.get('Job Number', '')} ({row.get('RT Hours', 0)}RT/{row.get('OT Hours', 0)}OT)"
                    )
                    key = str(actual_index)
                    option_labels[key] = label
                    option_keys.append(key)

                selected_multiple_delete = st.multiselect(
                    f"Select multiple entries to delete (for {date_val}):",
                    option_keys,
                    default=[],
                    placeholder="Choose multiple entries to delete...",
                    format_func=lambda opt: option_labels.get(opt, opt),
                    key=f"delete_multi_options_{selected_date_str}"
                )

                col1, col2 = st.columns([2, 1])
                with col2:
                    delete_multiple_button = st.button("Delete Selected Entries", type="secondary", key=f"delete_multiple_{selected_date_str}")

                if delete_multiple_button and selected_multiple_delete:
                    with st.spinner("Deleting selected entries..."):
                        try:
                            indices_to_delete = [int(opt) for opt in selected_multiple_delete]

                            if indices_to_delete:
                                updated_data = total_data.drop(index=indices_to_delete).reset_index(drop=True)
                                excel_synced = _write_time_data_to_excel(updated_data)
                                synced = _replace_time_data_in_google(updated_data)
                                if not excel_synced:
                                    st.warning("Deleted entries locally, but could not update local Time Data worksheet.")
                                if not synced:
                                    st.warning("Deleted entries locally, but external Time Data storage could not be updated.")

                                st.session_state.session_time_data = updated_data
                                st.session_state['sheet_cache_token'] = st.session_state.get('sheet_cache_token', 0) + 1
                                st.success(f"Deleted {len(indices_to_delete)} selected entries from {date_val}.")
                                st.rerun()
                            else:
                                st.error("No valid entries selected for deletion.")

                        except ValueError as e:
                            st.error(f"Could not parse row selections: {e}")
                        except Exception as e:
                            st.error(f"Could not delete entries: {e}")

            with delete_tabs[1]:
                st.warning(f"This will delete ALL {filtered_entries} entries for {date_val}")
                col1, col2 = st.columns([2, 1])
                with col2:
                    delete_all_button = st.button("Delete All Entries", type="secondary", key="delete_all")

                if delete_all_button:
                    with st.spinner("Deleting all entries..."):
                        try:
                            remaining_data = total_data[
                                pd.to_datetime(total_data["Date"]).dt.strftime("%Y-%m-%d") != selected_date_str
                            ].reset_index(drop=True) if not total_data.empty else pd.DataFrame()

                            excel_synced = _write_time_data_to_excel(remaining_data)
                            synced = _replace_time_data_in_google(remaining_data)
                            if not excel_synced:
                                st.warning("Deleted entries locally, but could not update local Time Data worksheet.")
                            if not synced:
                                st.warning("Deleted entries locally, but external Time Data storage could not be updated.")

                            st.session_state.session_time_data = remaining_data
                            st.session_state['sheet_cache_token'] = st.session_state.get('sheet_cache_token', 0) + 1
                            st.success(f"Deleted {filtered_entries} entries from {date_val}.")
                            st.rerun()

                        except Exception as e:
                            st.error(f"Could not delete entries: {e}")
        else:
            st.info("⚠️ Admin access required to delete entries.")
    else:
        st.info(f"No time data entries found for {date_val}")
        
except Exception as e:
    st.info(f"Time Data sheet not found or empty. Error: {e}")

# --- Export functionality - Admin only ---
if user_type.upper() == "ADMIN":
    st.divider()
    st.subheader("Export to Templates")

    try:
        time_data_for_export = get_time_data_from_session(None)
        if not time_data_for_export.empty:
            time_data_for_export = time_data_for_export.copy()
        else:
            time_data_for_export = smart_read_data("Time Data", force_refresh=True)
            if time_data_for_export.empty:
                time_data_for_export = safe_read_excel(XLSX, "Time Data")
        if not time_data_for_export.empty:

            def _prepare_employee_entries(entries):
                """Prepare employee entries for Daily Time export (don't combine, keep separate)"""
                if not entries:
                    return []

                prepared_entries = []
                for entry in entries:
                    prepared_entry = entry.copy()
                    prepared_entries.append(prepared_entry)

                return prepared_entries

            def _write_employee_to_daily_time(ws, emp_entries, row_num, employee_info, cost_code_descriptions):
                """Write employee data to specific row in Daily Time template"""
                if not emp_entries:
                    return 0

                emp_name = str(emp_entries[0].get('Name', ''))

                # Base employee info (columns A-D)
                emp_info = employee_info.get(emp_name, {})
                trade_class = str(emp_entries[0].get('Trade Class', '') or emp_info.get('override_trade_class', '') or '')
                ws.cell(row=row_num, column=1, value=emp_name)
                ws.cell(row=row_num, column=2, value=trade_class)

                rate_values = []
                for key in ('premium_rate', 'subsistence_rate', 'travel_rate'):
                    raw_val = emp_info.get(key, '')
                    if raw_val is None:
                        raw_val = ''
                    val = str(raw_val).strip()
                    if val.lower() == 'nan':
                        val = ''
                    if val:
                        rate_values.append(val)

                rate_value = rate_values[0] if rate_values else ''
                ws.cell(row=row_num, column=4, value=rate_value or None)

                # First entry goes in columns E-I
                if len(emp_entries) >= 1:
                    entry = emp_entries[0]
                    cost_code = str(entry.get('Cost Code', ''))
                    cost_desc = cost_code_descriptions.get(cost_code, '')

                    job_display = f"{entry.get('Job Number', '')} - {str(entry.get('Job Area', '')).zfill(3)} - {entry.get('Description of work', '')}"

                    ws.cell(row=row_num, column=5, value=cost_desc)  # E: Cost code description
                    ws.cell(row=row_num, column=6, value=cost_code)  # F: Cost code
                    ws.cell(row=row_num, column=7, value=job_display)  # G: Job Number - Area - Description
                    ws.cell(row=row_num, column=8, value=float(entry.get('RT Hours', 0) or 0))  # H: RT
                    ws.cell(row=row_num, column=9, value=float(entry.get('OT Hours', 0) or 0))   # I: OT

                # Second entry goes in columns K-O
                if len(emp_entries) >= 2:
                    entry = emp_entries[1]
                    cost_code = str(entry.get('Cost Code', ''))
                    cost_desc = cost_code_descriptions.get(cost_code, '')

                    job_display_2 = f"{entry.get('Job Number', '')} - {str(entry.get('Job Area', '')).zfill(3)} - {entry.get('Description of work', '')}"

                    ws.cell(row=row_num, column=11, value=cost_desc)  # K: Cost code description
                    ws.cell(row=row_num, column=12, value=cost_code)  # L: Cost code
                    ws.cell(row=row_num, column=13, value=job_display_2)  # M: Job Number - Area - Description
                    ws.cell(row=row_num, column=14, value=float(entry.get('RT Hours', 0) or 0))  # N: RT
                    ws.cell(row=row_num, column=15, value=float(entry.get('OT Hours', 0) or 0))   # O: OT

                # Return number of rows used (1 if 1-2 entries, more if 3+ entries)
                rows_used = 1
                if len(emp_entries) > 2:
                    # For 3+ entries, we need additional rows
                    additional_entries = emp_entries[2:]
                    current_extra_row = row_num + 1

                    for i, entry in enumerate(additional_entries):
                        if i % 2 == 0:  # First entry of new row (columns E-I)
                            cost_code = str(entry.get('Cost Code', ''))
                            cost_desc = cost_code_descriptions.get(cost_code, '')
                            job_display = f"{entry.get('Job Number', '')} - {str(entry.get('Job Area', '')).zfill(3)} - {entry.get('Description of work', '')}"

                            ws.cell(row=current_extra_row, column=1, value=emp_name)
                            ws.cell(row=current_extra_row, column=5, value=cost_desc)
                            ws.cell(row=current_extra_row, column=6, value=cost_code)
                            ws.cell(row=current_extra_row, column=7, value=job_display)
                            ws.cell(row=current_extra_row, column=8, value=float(entry.get('RT Hours', 0) or 0))
                            ws.cell(row=current_extra_row, column=9, value=float(entry.get('OT Hours', 0) or 0))

                        else:  # Second entry of row (columns K-O)
                            cost_code = str(entry.get('Cost Code', ''))
                            cost_desc = cost_code_descriptions.get(cost_code, '')
                            job_display = f"{entry.get('Job Number', '')} - {str(entry.get('Job Area', '')).zfill(3)} - {entry.get('Description of work', '')}"

                            ws.cell(row=current_extra_row, column=11, value=cost_desc)
                            ws.cell(row=current_extra_row, column=12, value=cost_code)
                            ws.cell(row=current_extra_row, column=13, value=job_display)
                            ws.cell(row=current_extra_row, column=14, value=float(entry.get('RT Hours', 0) or 0))
                            ws.cell(row=current_extra_row, column=15, value=float(entry.get('OT Hours', 0) or 0))
                            current_extra_row += 1

                        if i % 2 == 0 and i == len(additional_entries) - 1:
                            # If we have an odd number of additional entries, move to next row
                            current_extra_row += 1

                    rows_used = current_extra_row - row_num

                return rows_used

            def create_template_exports(export_date):
                """Create proper template-based exports using Daily Time.xlsx and TimeEntries.xlsx"""
                filtered_data = time_data_for_export[
                    pd.to_datetime(time_data_for_export['Date']).dt.date == export_date
                ].copy()

                if filtered_data.empty:
                    st.warning(f"No data found for {export_date}")
                    return None

                try:
                    zip_buffer = io.BytesIO()

                    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    
                        daily_time_template = Path(__file__).resolve().parent.parent / "Daily Time.xlsx"
                        if daily_time_template.exists():
                            tmp_path = None
                            try:
                                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                                    tmp_path = tmp.name
                                    copyfile(daily_time_template, tmp_path)

                                wb = load_workbook(tmp_path)
                                ws = wb.active

                                for row_idx in range(1, 15):
                                    for col_idx in range(1, 15):
                                        cell = ws.cell(row=row_idx, column=col_idx)
                                        if cell.value and 'DATA DATE' in str(cell.value).upper():
                                            ws.cell(row=row_idx, column=col_idx + 1, value=export_date.strftime('%Y-%m-%d'))
                                            break

                                # Load employee data to determine indirect/direct status
                                employee_df = safe_read_excel(XLSX, "Employee List")
                                employee_info = {}
                                if not employee_df.empty:
                                    for _, emp_row in employee_df.iterrows():
                                        name = str(emp_row.get("Employee Name", ""))
                                        employee_info[name] = {
                                            'indirect': str(emp_row.get("Indirect / Direct", "")).strip().upper() == "INDIRECT",
                                            'override_trade_class': str(emp_row.get("Override Trade Class", "") or ""),
                                            'premium_rate': str(emp_row.get("Premium Rate", "") or ""),
                                            'subsistence_rate': str(emp_row.get("Subsistence Rate", "") or ""),
                                            'travel_rate': str(emp_row.get("Travel Rate", "") or ""),
                                            'time_record_type': str(emp_row.get("Time Record Type", "") or "").strip()
                                        }

                                # Load cost codes for descriptions
                                cost_codes_df = safe_read_excel(XLSX, "Cost Codes")
                                cost_code_descriptions = {}
                                if not cost_codes_df.empty:
                                    for _, cc_row in cost_codes_df.iterrows():
                                        code = str(cc_row.get("Cost Code", "") or "").strip()
                                        desc = str(cc_row.get("Description", "") or "").strip()
                                        if code:
                                            cost_code_descriptions[code] = desc

                                # Group data by employee and combine entries for same employee
                                employee_groups = {}
                                for _, row in filtered_data.iterrows():
                                    emp_name = str(row.get('Name', ''))
                                    if emp_name not in employee_groups:
                                        employee_groups[emp_name] = []
                                    employee_groups[emp_name].append(row)

                                # Separate indirect and direct employees
                                indirect_employees = []
                                direct_employees = []

                                for emp_name, entries in employee_groups.items():
                                    is_indirect = employee_info.get(emp_name, {}).get('indirect', False)
                                    prepared_entries = _prepare_employee_entries(entries)

                                    employee_data = {
                                        'name': emp_name,
                                        'entries': prepared_entries,
                                        'is_indirect': is_indirect
                                    }

                                    if is_indirect:
                                        indirect_employees.append(employee_data)
                                    else:
                                        direct_employees.append(employee_data)

                                # Place indirect employees in rows 8-30
                                current_row = 8
                                used_indirect_rows = []
                                for emp_data in indirect_employees:
                                    if current_row > 30:
                                        break
                                    rows_used = _write_employee_to_daily_time(ws, emp_data['entries'], current_row, employee_info, cost_code_descriptions)
                                    for i in range(rows_used):
                                        used_indirect_rows.append(current_row + i)
                                    current_row += rows_used

                                # Place direct employees in rows 32-261
                                current_row = 32
                                used_direct_rows = []
                                for emp_data in direct_employees:
                                    if current_row > 261:
                                        break
                                    rows_used = _write_employee_to_daily_time(ws, emp_data['entries'], current_row, employee_info, cost_code_descriptions)
                                    for i in range(rows_used):
                                        used_direct_rows.append(current_row + i)
                                    current_row += rows_used

                                # Hide unused rows
                                # Hide unused indirect rows (8-30)
                                for row_num in range(8, 31):
                                    if row_num not in used_indirect_rows:
                                        ws.row_dimensions[row_num].hidden = True

                                # Hide unused direct rows (32-261)
                                for row_num in range(32, 262):
                                    if row_num not in used_direct_rows:
                                        ws.row_dimensions[row_num].hidden = True

                                # Add job summaries starting at row 264
                                current_summary_row = 264

                                # Group data by job number, area, and description
                                job_groups = {}
                                for _, row in filtered_data.iterrows():
                                    job_num = str(row.get('Job Number', ''))
                                    job_area = str(row.get('Job Area', '')).zfill(3)
                                    job_desc = str(row.get('Description of work', ''))
                                    job_key = f"{job_num} - {job_area} - {job_desc}"

                                    if job_key not in job_groups:
                                        job_groups[job_key] = set()

                                    comment = str(row.get('Comments', '')).strip()
                                    if comment and comment.lower() not in ['nan', 'none', '']:
                                        job_groups[job_key].add(comment)

                                # Write job summaries (only if there are comments)
                                for job_key, comments in job_groups.items():
                                    if current_summary_row > 500:  # Avoid going too far down
                                        break

                                    # Only write if there are actual comments
                                    if comments:
                                        # Write job header (bold and underlined)
                                        cell = ws.cell(row=current_summary_row, column=1, value=job_key)
                                        cell.font = openpyxl.styles.Font(bold=True, underline='single')
                                        current_summary_row += 1

                                        # Write each unique comment
                                        for comment in sorted(comments):
                                            if current_summary_row > 500:
                                                break
                                            ws.cell(row=current_summary_row, column=1, value=comment)
                                            current_summary_row += 1

                                        # Add blank line between job groups
                                        current_summary_row += 1

                                wb.save(tmp_path)
                                wb.close()

                                with open(tmp_path, 'rb') as f:
                                    zip_file.writestr(f"{export_date.strftime('%m-%d-%Y')} - Daily Time.xlsx", f.read())

                            finally:
                                if tmp_path and os.path.exists(tmp_path):
                                    try:
                                        os.unlink(tmp_path)
                                    except Exception:
                                        pass
                        
                        timeentries_template = Path(__file__).resolve().parent.parent / "TimeEntries.xlsx"
                        if timeentries_template.exists():
                            # Load employee data for rates (if not already loaded)
                            if 'employee_info' not in locals():
                                employee_df = smart_read_data("Employee List", force_refresh=True)
                                employee_info = {}
                                if isinstance(employee_df, pd.DataFrame) and not employee_df.empty:
                                    employee_df = employee_df.copy()
                                    employee_df.columns = [str(c).strip() for c in employee_df.columns]
                                    for _, emp_row in employee_df.iterrows():
                                        name = str(emp_row.get("Employee Name", "")).strip()
                                        if not name:
                                            continue
                                        employee_info[name] = {
                                            'indirect': str(emp_row.get("Indirect / Direct", "")).strip().upper() == "INDIRECT",
                                            'premium_rate': str(emp_row.get("Premium Rate", "") or "").strip(),
                                            'subsistence': str(emp_row.get("Subsistence Rate", "") or "").strip(),
                                            'travel_rate': str(emp_row.get("Travel Rate", "") or "").strip(),
                                            'night_shift': _normalize_night_flag(emp_row.get("Night Shift", "")),
                                            'time_record_type': str(emp_row.get("Time Record Type", "") or "").strip()
                                        }

                            unique_jobs = filtered_data['Job Number'].dropna().unique()

                            for job in unique_jobs:
                                if pd.isna(job) or str(job).strip() == '':
                                    continue
                                
                                job_data = filtered_data[filtered_data['Job Number'].astype(str).str.strip() == str(job).strip()].copy()
                                if job_data.empty:
                                    continue

                                tmp_path = None
                                try:
                                    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                                        tmp_path = tmp.name
                                        copyfile(timeentries_template, tmp_path)

                                    wb = load_workbook(tmp_path)
                                    ws = wb.active

                                    current_row = 4
                                    for _, row in job_data.iterrows():
                                        # Get employee data for rates
                                        emp_name = str(row.get('Name', ''))
                                        emp_info = employee_info.get(emp_name, {})

                                        # Helper function to clean values
                                        def clean_value(val):
                                            if pd.isna(val) or str(val).lower() in ['nan', 'none', '']:
                                                return ''
                                            return str(val)

                                        premium_rate = clean_value(emp_info.get('premium_rate', '')) or clean_value(row.get('Premium Rate', ''))
                                        subsistence_rate = clean_value(emp_info.get('subsistence', '')) or clean_value(row.get('Subsistence Rate', ''))
                                        travel_rate = clean_value(emp_info.get('travel_rate', '')) or clean_value(row.get('Travel Rate', ''))
                                        night_shift = _normalize_night_flag(emp_info.get('night_shift', ''))
                                        if not night_shift:
                                            night_shift = _normalize_night_flag(row.get('Night Shift', ''))

                                        time_record_type = clean_value(emp_info.get('time_record_type', ''))

                                        rate_cell = subsistence_rate or premium_rate or travel_rate

                                        base_data = [
                                            export_date.strftime('%Y-%m-%d'),  # A - Date
                                            time_record_type,                  # B - Time Record Type
                                            clean_value(row.get('Employee Number', '')),  # C - Employee Number
                                            clean_value(row.get('Name', '')),             # D - Name
                                            clean_value(row.get('Trade Class', '')),      # E - Trade Class
                                            'Y',                               # F - Always 'Y'
                                            clean_value(row.get('Cost Code', '')),        # G - Cost Code
                                            str(row.get('Job Area', '')).zfill(3),        # H - Job Area (3 digits)
                                            '',                                # I - Empty
                                            '211',                             # J - Pay Code (will be changed per entry)
                                            0.0,                               # K - Hours (will be set per entry)
                                            night_shift,                       # L - Night Shift
                                            rate_cell,                         # M - Rate cell (prefers Subsistence)
                                            '',                                # N - Comments (left blank)
                                            ''                                 # O - Extra column (left blank)
                                        ]

                                        rt_hours = float(row.get('RT Hours', 0) or 0)
                                        if rt_hours > 0:
                                            rt_data = base_data.copy()
                                            rt_data[9] = '211'  # Regular time pay code
                                            rt_data[10] = rt_hours

                                            for col, val in enumerate(rt_data, 1):
                                                ws.cell(row=current_row, column=col, value=val)
                                            current_row += 1

                                        ot_hours = float(row.get('OT Hours', 0) or 0)
                                        if ot_hours > 0:
                                            ot_data = base_data.copy()
                                            ot_data[9] = '212'  # Overtime pay code
                                            ot_data[10] = ot_hours

                                            for col, val in enumerate(ot_data, 1):
                                                ws.cell(row=current_row, column=col, value=val)
                                            current_row += 1

                                        # Add subsistence entry if employee has subsistence rate
                                        if subsistence_rate and subsistence_rate != '':
                                            try:
                                                subsistence_amount = float(subsistence_rate)
                                                if subsistence_amount > 0:
                                                    sub_data = base_data.copy()
                                                    sub_data[9] = '261'  # Subsistence pay code
                                                    sub_data[10] = 1.0   # Hours = 1 for subsistence
                                                    sub_data[12] = subsistence_rate  # Put subsistence rate in column M (Premium Rate column)

                                                    for col, val in enumerate(sub_data, 1):
                                                        ws.cell(row=current_row, column=col, value=val)
                                                    current_row += 1
                                            except (ValueError, TypeError):
                                                pass  # Skip if subsistence rate is not a valid number

                                    wb.save(tmp_path)
                                    wb.close()

                                    with open(tmp_path, 'rb') as f:
                                        zip_file.writestr(f"{export_date.strftime('%m-%d-%Y')} - {job} - Daily Import.xlsx", f.read())

                                finally:
                                    if tmp_path and os.path.exists(tmp_path):
                                        try:
                                            os.unlink(tmp_path)
                                        except Exception:
                                            pass
                    
                    zip_buffer.seek(0)
                    return zip_buffer.getvalue()
                    
                except Exception as e:
                    st.error(f"Error creating export: {e}")
                    return None

            if st.button("Create & Download Exports", type="primary"):
                with st.spinner("Creating export package..."):
                    zip_data = create_template_exports(date_val)
                    if zip_data:
                        st.download_button(
                            label="Download Export Package",
                            data=zip_data,
                            file_name=f"Timesheet_Export_{date_val.strftime('%m-%d-%Y')}.zip",
                            mime="application/zip",
                            key="download_exports"
                        )
                        st.success(f"Export package ready for download! Contains Daily Time and Daily Import files for {date_val}")
                    else:
                        st.error("Failed to create export package")

        else:
            st.info("No time data available for export. Add some entries first.")

    except Exception:
        st.info("No time data available for export. Add some entries first.")

else:
    st.divider()
    st.info("⚠️ Admin access required to export templates.")


