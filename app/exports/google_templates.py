from __future__ import annotations

import io
import base64
import time
from datetime import date, datetime
from typing import Iterable

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.integrations.google_sheets import get_sheets_manager


def _normalize_name(value: object) -> str:
    return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum())


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    text = str(value).strip()
    return "" if text.lower() in {"nan", "none"} else text


def _is_truthy(value: object) -> bool:
    if isinstance(value, bool):
        return value
    text = _clean_text(value).lower()
    if text in {"true", "yes", "y", "1", "active", "enabled"}:
        return True
    try:
        return float(text) == 1.0
    except Exception:
        return False


def _find_column(df: pd.DataFrame, candidates: Iterable[str], fallback_index: int | None = None) -> str | None:
    if df is None or df.empty:
        return None
    columns = [str(col).strip() for col in df.columns]
    exact = {col: col for col in columns}
    for candidate in candidates:
        if candidate in exact:
            return exact[candidate]
    normalized = {_normalize_name(col): col for col in columns}
    for candidate in candidates:
        match = normalized.get(_normalize_name(candidate))
        if match:
            return match
    if fallback_index is not None and len(columns) > fallback_index:
        return columns[fallback_index]
    return None


def get_google_template_workbook_bytes(spreadsheet_id: str | None = None) -> bytes:
    """Return an Excel export of the configured Google Sheets workbook."""
    sheet_id = str(spreadsheet_id or st.secrets.get("google_sheets_id", "")).strip()
    if not sheet_id:
        raise RuntimeError("Google Sheets ID is not configured.")
    manager = get_sheets_manager()
    exporter = getattr(manager, "export_spreadsheet_xlsx", None)
    if callable(exporter):
        return exporter(sheet_id)

    session_getter = getattr(manager, "_ensure_session", None)
    if not callable(session_getter):
        raise RuntimeError("Google Sheets connection is not configured.")
    session = session_getter()
    if session is None:
        raise RuntimeError("Google Sheets connection is not configured.")

    url = f"https://www.googleapis.com/drive/v3/files/{sheet_id}/export"
    params = {
        "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    response = session.get(url, params=params)
    response.raise_for_status()
    return response.content


def load_template_sheet_workbook(template_bytes: bytes, sheet_candidates: Iterable[str]) -> tuple[Workbook, Worksheet]:
    """Load one worksheet from a Google-exported workbook and remove the other tabs."""
    wb = load_workbook(io.BytesIO(template_bytes))
    candidates = list(sheet_candidates)
    normalized_candidates = {_normalize_name(candidate) for candidate in candidates}

    selected_name = None
    sheet_lookup = {_normalize_name(name): name for name in wb.sheetnames}
    for normalized_candidate in normalized_candidates:
        if normalized_candidate in sheet_lookup:
            selected_name = sheet_lookup[normalized_candidate]
            break

    if selected_name is None:
        for name in wb.sheetnames:
            normalized_sheet = _normalize_name(name)
            if any(candidate and candidate in normalized_sheet for candidate in normalized_candidates):
                selected_name = name
                break

    if selected_name is None:
        wanted = ", ".join(candidates)
        found = ", ".join(wb.sheetnames)
        raise RuntimeError(f"Template worksheet not found. Wanted one of: {wanted}. Found: {found}")

    ws = wb[selected_name]
    wb.active = wb.index(ws)
    for sheet_name in list(wb.sheetnames):
        if sheet_name != selected_name:
            del wb[sheet_name]

    if ws.sheet_view:
        ws.sheet_view.topLeftCell = "A1"
        if ws.sheet_view.selection:
            ws.sheet_view.selection[0].activeCell = "A1"
            ws.sheet_view.selection[0].sqref = "A1"
    return wb, ws


def workbook_to_bytes(wb: Workbook) -> bytes:
    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()
    buffer.seek(0)
    return buffer.getvalue()


def _a1_sheet_name(title: str) -> str:
    return "'" + str(title).replace("'", "''") + "'"


def _sign_in_employee_rows(employees_df: pd.DataFrame) -> tuple[list[list[str]], int, int | None]:
    if employees_df is None:
        employees_df = pd.DataFrame()
    employees_df = employees_df.copy()
    employees_df.columns = [str(col).strip() for col in employees_df.columns]

    company_col = _find_column(employees_df, ("Company Name", "Company", "Employer"), 11)
    name_col = _find_column(employees_df, ("Employee Name", "Name", "Employee"), 2)
    craft_col = _find_column(employees_df, ("Craft / Certification", "Craft Certification", "Craft", "Certification"), 12)
    active_col = _find_column(employees_df, ("Active", "Is Active", "Enabled"), 13)

    if active_col:
        employees_df = employees_df[employees_df[active_col].apply(_is_truthy)]

    rows: list[list[str]] = []
    for _, employee in employees_df.iterrows():
        if len(rows) >= 64:
            break
        rows.append([
            _clean_text(employee.get(company_col, "")) if company_col else "",
            _clean_text(employee.get(name_col, "")) if name_col else "",
            _clean_text(employee.get(craft_col, "")) if craft_col else "",
        ])
    active_count = len(rows)
    rows.extend([["", "", ""] for _ in range(max(64 - len(rows), 0))])
    first_hidden_row = 11 + active_count + 5
    if first_hidden_row > 74:
        first_hidden_row = None
    return rows, active_count, first_hidden_row


def _sign_in_client_rows(clients_df: pd.DataFrame) -> tuple[list[list[str]], int, int | None]:
    if clients_df is None:
        clients_df = pd.DataFrame()
    clients_df = clients_df.copy()
    clients_df.columns = [str(col).strip() for col in clients_df.columns]

    company_col = _find_column(clients_df, ("COMPANY", "Company Name", "Company", "Client Company"), 0)
    name_col = _find_column(clients_df, ("PERSON NAME", "Person Name", "Client Name", "Name"), 1)
    cert_col = _find_column(clients_df, ("CERTIFICATION", "Certification", "Craft / Certification", "Craft"), 2)
    active_col = _find_column(clients_df, ("Active", "Is Active", "Enabled"), 3)

    if active_col:
        clients_df = clients_df[clients_df[active_col].apply(_is_truthy)]

    rows: list[list[str]] = []
    for _, client in clients_df.iterrows():
        if len(rows) >= 18:
            break
        rows.append([
            _clean_text(client.get(company_col, "")) if company_col else "",
            _clean_text(client.get(name_col, "")) if name_col else "",
            _clean_text(client.get(cert_col, "")) if cert_col else "",
        ])
    active_count = len(rows)
    rows.extend([["", "", ""] for _ in range(max(18 - len(rows), 0))])
    first_hidden_row = 76 + active_count + 3
    if first_hidden_row > 93:
        first_hidden_row = None
    return rows, active_count, first_hidden_row


def _find_sheet_id(metadata: dict, title: str) -> int:
    target = _normalize_name(title)
    for sheet in metadata.get("sheets", []):
        properties = sheet.get("properties", {})
        if _normalize_name(properties.get("title")) == target:
            return int(properties["sheetId"])
    found = ", ".join(
        str(sheet.get("properties", {}).get("title", ""))
        for sheet in metadata.get("sheets", [])
    )
    raise RuntimeError(f"Worksheet '{title}' not found. Found: {found}")


def _merge_pdf_bytes(pdf_parts: list[bytes]) -> bytes:
    if not pdf_parts:
        raise RuntimeError("No sign in sheets were exported.")
    if len(pdf_parts) == 1:
        return pdf_parts[0]
    try:
        from pypdf import PdfReader, PdfWriter
    except ImportError as exc:
        raise RuntimeError("PDF merge support is not installed. Add pypdf to requirements.txt.") from exc

    writer = PdfWriter()
    for pdf_part in pdf_parts:
        reader = PdfReader(io.BytesIO(pdf_part))
        for page in reader.pages:
            writer.add_page(page)
    out = io.BytesIO()
    writer.write(out)
    out.seek(0)
    return out.getvalue()


def build_pdf_image_print_html(pdf_bytes: bytes, *, auto_print: bool = True) -> str:
    """Render PDF pages as browser-printable images and open the print dialog."""
    try:
        import fitz
    except ImportError as exc:
        raise RuntimeError("PDF print rendering support is not installed. Add PyMuPDF to requirements.txt.") from exc

    image_tags: list[str] = []
    document = fitz.open(stream=pdf_bytes, filetype="pdf")
    try:
        for page in document:
            pixmap = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
            encoded_image = base64.b64encode(pixmap.tobytes("png")).decode("ascii")
            image_tags.append(
                f'<section class="print-page"><img src="data:image/png;base64,{encoded_image}" alt="Sign In Sheet page"></section>'
            )
    finally:
        document.close()

    if not image_tags:
        raise RuntimeError("No printable pages were found in the Sign In Sheet PDF.")

    pages_html = "\n".join(image_tags)
    auto_print_script = ""
    if auto_print:
        auto_print_script = """
            window.addEventListener("load", function () {
                setTimeout(openPrintDialog, 500);
            });
        """
    return f"""
    <!doctype html>
    <html>
    <head>
    <meta charset="utf-8">
    <style>
        body {{
            margin: 0;
            font-family: Arial, Helvetica, sans-serif;
            background: #ffffff;
            color: #111827;
        }}
        .controls {{
            display: flex;
            align-items: center;
            gap: 10px;
            padding: 10px 12px;
            border: 1px solid #d1d5db;
            background: #ffffff;
        }}
        button, a {{
            border: 1px solid #9ca3af;
            background: #ffffff;
            color: #111827;
            padding: 6px 10px;
            font-size: 14px;
            text-decoration: none;
            cursor: pointer;
        }}
        .print-pages {{
            background: #ffffff;
        }}
        .print-page {{
            width: 100%;
            min-height: 100vh;
            display: flex;
            align-items: flex-start;
            justify-content: center;
            background: #ffffff;
            page-break-after: always;
        }}
        .print-page:last-child {{
            page-break-after: auto;
        }}
        .print-page img {{
            display: block;
            width: 100%;
            height: auto;
        }}
        @page {{
            size: letter portrait;
            margin: 0;
        }}
        @media print {{
            body {{
                background: #ffffff;
            }}
            .controls {{
                display: none;
            }}
            .print-page {{
                width: 100vw;
                height: 100vh;
                min-height: 100vh;
                overflow: hidden;
                page-break-after: always;
            }}
            .print-page:last-child {{
                page-break-after: auto;
            }}
            .print-page img {{
                width: 100%;
                height: 100%;
                object-fit: contain;
            }}
        }}
    </style>
    </head>
    <body>
    <div class="controls">
        <span id="status">The print dialog should open automatically.</span>
        <button onclick="openPrintDialog()">Open print dialog</button>
    </div>
    <main class="print-pages">
        {pages_html}
    </main>
    <script>
        const status = document.getElementById("status");
        function openPrintDialog() {{
            try {{
                window.focus();
                window.print();
            }} catch (error) {{
                status.textContent = "Use your browser print command to print these sheets.";
            }}
        }}
        {auto_print_script}
    </script>
    </body>
    </html>
    """


def build_sign_in_sheet_pdf(
    employees_df: pd.DataFrame,
    sign_in_dates: Iterable[date | datetime],
    clients_df: pd.DataFrame | None = None,
    spreadsheet_id: str | None = None,
) -> tuple[bytes, int, int, int]:
    """Create Google-rendered PDFs from temporary copies of the Sign In Sheet tab."""
    dates = list(sign_in_dates)
    if not dates:
        raise ValueError("At least one sign in date is required.")

    sheet_id = str(spreadsheet_id or st.secrets.get("google_sheets_id", "")).strip()
    if not sheet_id:
        raise RuntimeError("Google Sheets ID is not configured.")

    rows, active_count, first_hidden_row = _sign_in_employee_rows(employees_df)
    client_rows, active_client_count, first_hidden_client_row = _sign_in_client_rows(clients_df)
    manager = get_sheets_manager()
    metadata = manager.get_spreadsheet_metadata(sheet_id, fields="sheets(properties(sheetId,title,index))")
    source_sheet_id = _find_sheet_id(metadata, "Sign In Sheet")

    created_sheet_ids: list[int] = []
    pdf_parts: list[bytes] = []
    unique_token = int(time.time() * 1000)

    try:
        duplicate_requests = []
        normalized_dates: list[date] = []
        for index, sign_in_date in enumerate(dates):
            if isinstance(sign_in_date, datetime):
                sign_in_date = sign_in_date.date()
            normalized_dates.append(sign_in_date)
            temp_title = f"_Print Sign In {sign_in_date:%Y-%m-%d} {unique_token}-{index + 1}"
            duplicate_requests.append({
                "duplicateSheet": {
                    "sourceSheetId": source_sheet_id,
                    "newSheetName": temp_title,
                }
            })

        duplicate_response = manager.batch_update(sheet_id, duplicate_requests)
        temp_sheets: list[tuple[date, int, str]] = []
        for index, sign_in_date in enumerate(normalized_dates):
            duplicate_properties = duplicate_response["replies"][index]["duplicateSheet"]["properties"]
            temp_sheet_id = int(duplicate_properties["sheetId"])
            temp_sheet_title = str(duplicate_properties["title"])
            created_sheet_ids.append(temp_sheet_id)
            temp_sheets.append((sign_in_date, temp_sheet_id, temp_sheet_title))

        value_updates = []
        for sign_in_date, _, temp_sheet_title in temp_sheets:
            sheet_name = _a1_sheet_name(temp_sheet_title)
            value_updates.extend([
                {
                    "range": f"{sheet_name}!D6",
                    "values": [[sign_in_date.strftime("%Y/%m/%d")]],
                },
                {
                    "range": f"{sheet_name}!A11:C74",
                    "values": rows,
                },
                {
                    "range": f"{sheet_name}!A76:C93",
                    "values": client_rows,
                },
            ])

        batch_value_updater = getattr(manager, "batch_update_values", None)
        if callable(batch_value_updater):
            batch_value_updater(sheet_id, value_updates)
        else:
            for value_update in value_updates:
                manager.update_values(sheet_id, value_update["range"], value_update["values"])

        hide_requests = []
        for _, temp_sheet_id, _ in temp_sheets:
            hide_requests.extend([{
                "updateSheetProperties": {
                    "properties": {
                        "sheetId": temp_sheet_id,
                        "gridProperties": {
                            "frozenRowCount": 10,
                        },
                    },
                    "fields": "gridProperties.frozenRowCount",
                }
            }, {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": temp_sheet_id,
                        "dimension": "ROWS",
                        "startIndex": 10,
                        "endIndex": 74,
                    },
                    "properties": {"hiddenByUser": False},
                    "fields": "hiddenByUser",
                }
            }, {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": temp_sheet_id,
                        "dimension": "ROWS",
                        "startIndex": 75,
                        "endIndex": 93,
                    },
                    "properties": {"hiddenByUser": False},
                    "fields": "hiddenByUser",
                }
            }])
            if first_hidden_row is not None:
                hide_requests.append({
                    "updateDimensionProperties": {
                        "range": {
                            "sheetId": temp_sheet_id,
                            "dimension": "ROWS",
                            "startIndex": first_hidden_row - 1,
                            "endIndex": 74,
                        },
                        "properties": {"hiddenByUser": True},
                        "fields": "hiddenByUser",
                    }
                })
            if first_hidden_client_row is not None:
                hide_requests.append({
                    "updateDimensionProperties": {
                        "range": {
                            "sheetId": temp_sheet_id,
                            "dimension": "ROWS",
                            "startIndex": first_hidden_client_row - 1,
                            "endIndex": 93,
                        },
                        "properties": {"hiddenByUser": True},
                        "fields": "hiddenByUser",
                    }
                })
        manager.batch_update(sheet_id, hide_requests)
        time.sleep(0.35)
        narrow_margins = {"top": 0.25, "bottom": 0.25, "left": 0.25, "right": 0.25}
        for _, temp_sheet_id, _ in temp_sheets:
            pdf_parts.append(
                manager.export_sheet_pdf(
                    sheet_id,
                    temp_sheet_id,
                    repeat_frozen_rows=True,
                    margins=narrow_margins,
                )
            )
    finally:
        if created_sheet_ids:
            delete_requests = [{"deleteSheet": {"sheetId": temp_sheet_id}} for temp_sheet_id in created_sheet_ids]
            try:
                manager.batch_update(sheet_id, delete_requests)
            except Exception:
                pass

    return _merge_pdf_bytes(pdf_parts), active_count, active_client_count, len(dates)
