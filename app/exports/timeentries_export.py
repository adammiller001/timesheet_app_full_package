import io
import pandas as pd
from datetime import date
from openpyxl import load_workbook
from app.config import APP_DIR
from app.data.workbook import get_time_data, pad_job_area
from app.utils.excel_style import clone_row_styles

EXPECTED_HEADERS = ['Date','Time Record Type','Person Number','Employee Name','Override Trade Class','Post To Payroll','Cost Code / Phase','JobArea','Scope Change','Pay Code','Hours','Night Shift','Premium Rate / Subsistence Rate / Travel Rate','Comments']
TEMPLATE_EXPORT_BOOK = APP_DIR.parent / "TimeEntries.xlsx"
PAYCODE_MAP = {"REG":"211","OT":"212","SUBSISTENCE":"261"}

def _build_rows(sub: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, r in sub.iterrows():
        reg_h = float(r.get("RT Hours",0) or 0.0)
        ot_h  = float(r.get("OT Hours",0) or 0.0)
        base = {
            "Date": pd.to_datetime(r.get("Date","")).strftime("%Y-%m-%d"),
            "Time Record Type": "",
            "Person Number": r.get("Employee Number",""),
            "Employee Name": r.get("Name",""),
            "Override Trade Class": r.get("Trade Class",""),
            "Post To Payroll": "Y",
            "Cost Code / Phase": r.get("Class Type",""),
            "JobArea": pad_job_area(r.get("Job Area","")),
            "Scope Change": "",
            "Pay Code": "",
            "Hours": 0.0,
            "Night Shift": "",
            "Premium Rate / Subsistence Rate / Travel Rate": r.get("Premium Rate / Subsistence Rate / Travel Rate",""),
            "Comments": "",
        }
        if reg_h>0:
            t=base.copy(); t["Pay Code"]=PAYCODE_MAP.get("REG","211"); t["Hours"]=reg_h; rows.append(t)
        if ot_h>0:
            t=base.copy(); t["Pay Code"]=PAYCODE_MAP.get("OT","212");  t["Hours"]=ot_h; rows.append(t)
    return pd.DataFrame(rows, columns=EXPECTED_HEADERS)

def _find_template_sheet(wb):
    if "TimeEntries" in wb.sheetnames:
        return wb["TimeEntries"]
    for name in wb.sheetnames:
        if "timeentries" in name.lower():
            return wb[name]
    for name in wb.sheetnames:
        ws = wb[name]
        headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if headers[:len(EXPECTED_HEADERS)] == EXPECTED_HEADERS:
            return ws
    raise RuntimeError(f"Template workbook does not contain a compatible sheet. Found sheets: {wb.sheetnames}")

def _render_job(day_df: pd.DataFrame, job: str) -> bytes:
    subset = day_df[day_df["Job Number"].astype(str).str.strip() == str(job)].copy()
    out_df = _build_rows(subset)
    if not TEMPLATE_EXPORT_BOOK.exists():
        raise RuntimeError("Export template 'TimeEntries.xlsx' not found beside the app.")
    wb = load_workbook(TEMPLATE_EXPORT_BOOK)
    ws = _find_template_sheet(wb)
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    max_col = len(headers)
    data_start = 2
    has_template_data_row = ws.max_row >= 2
    for ridx, row in enumerate(out_df.itertuples(index=False), start=data_start):
        if has_template_data_row and ridx != 2:
            clone_row_styles(ws, ws, 2, ridx, max_col)
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=ridx, column=c_idx, value=val)
    last_written = data_start + len(out_df) - 1
    if has_template_data_row and ws.max_row > last_written:
        for r in range(last_written+1, ws.max_row+1):
            for c in range(1, max_col+1):
                ws.cell(row=r, column=c, value=None)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.getvalue()

def per_job_exports(xlsx_path: str, export_date: date):
    td = get_time_data(xlsx_path)
    if td.empty or "Date" not in td.columns:
        return []
    dmask = td["Date"].astype(str).str[:10] == export_date.strftime("%Y-%m-%d")
    day_df = td[dmask].copy()
    if day_df.empty:
        return []
    jobs_for_day = sorted(day_df["Job Number"].astype(str).str.strip().unique().tolist())
    for job in jobs_for_day:
        content = _render_job(day_df, job)
        file_name = f"{export_date.strftime('%m-%d-%Y')} - {job} - Daily Time Import.xlsx"
        yield file_name, content
