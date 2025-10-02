from __future__ import annotations

from pathlib import Path
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from typing import List

from app.config import get_default_xlsx_path
from app.integrations.google_sheets import get_sheets_manager

TIME_DATA_HEADERS: List[str] = [
    "Job Number",
    "Job Area",
    "Date",
    "Name",
    "Class Type",
    "Trade Class",
    "Employee Number",
    "RT Hours",
    "OT Hours",
    "Night Shift",
    "Premium Rate / Subsistence Rate / Travel Rate",
    "Comments",
]


def _clean_headers(df: pd.DataFrame) -> pd.DataFrame:
    try:
        df.columns = [str(c).strip() for c in df.columns]
    except Exception:
        pass
    return df


def _resolve_workbook_path(path_hint: str | Path | None = None) -> Path | None:
    candidate = path_hint or get_default_xlsx_path()
    if not candidate:
        return None
    try:
        path = Path(candidate)
    except TypeError:
        return None
    return path if path.exists() else None


def _read_local_sheet(
    possible_names: List[str],
    empty_cols: List[str] | None = None,
    workbook_path: str | Path | None = None,
) -> pd.DataFrame:
    path = _resolve_workbook_path(workbook_path)
    if path is None:
        return pd.DataFrame(columns=empty_cols or [])
    try:
        excel_file = pd.ExcelFile(path)
    except Exception:
        return pd.DataFrame(columns=empty_cols or [])

    def _match(name: str) -> str | None:
        target = name.strip().lower()
        for sheet in excel_file.sheet_names:
            if sheet.strip().lower() == target:
                return sheet
        return None

    for candidate in possible_names:
        sheet_name = _match(candidate)
        if not sheet_name:
            continue
        try:
            df = pd.read_excel(path, sheet_name=sheet_name)
            return _clean_headers(df)
        except Exception:
            continue
    return pd.DataFrame(columns=empty_cols or [])


def _read_sheet(
    possible_names: List[str],
    empty_cols: List[str] | None = None,
    force_refresh: bool = False,
    workbook_path: str | Path | None = None,
) -> pd.DataFrame:
    sheet_id = st.secrets.get("google_sheets_id", "")
    manager = get_sheets_manager()
    if force_refresh and hasattr(manager, "_data_cache"):
        manager._data_cache.pop(possible_names[0], None)  # type: ignore[attr-defined]

    if sheet_id:
        worksheet, actual_title = manager.find_worksheet(possible_names, sheet_id)
        if worksheet or actual_title:
            try:
                df = manager.read_worksheet(actual_title or possible_names[0], sheet_id, force_refresh=force_refresh)
                if isinstance(df, pd.DataFrame) and not df.empty:
                    return _clean_headers(df)
            except Exception:
                pass

    # Fallback to local workbook when available
    return _read_local_sheet(possible_names, empty_cols=empty_cols, workbook_path=workbook_path)


def get_employees(path: str | None = None, force_refresh: bool = False) -> pd.DataFrame:
    df = _read_sheet(["Employee List", "Employees"], ["Employee Name", "Employee Number", "Override Trade Class"], force_refresh, path)
    if not df.empty:
        df = df.rename(columns={
            "Employee Name": "name",
            "Employee Number": "emp_num",
            "Override Trade Class": "trade",
        })
    return df


def get_jobs(path: str | None = None, force_refresh: bool = False) -> pd.DataFrame:
    df = _read_sheet(["Job Numbers", "Jobs"], ["JOB #", "AREA #", "DESCRIPTION"], force_refresh, path)
    if not df.empty:
        df = df.rename(columns={
            "JOB #": "job_num",
            "AREA #": "area_code",
            "DESCRIPTION": "area_desc",
        })
        if "area_code" in df.columns:
            df["area_code"] = df["area_code"].apply(pad_job_area)
    return df


def get_cost_codes(path: str | None = None, force_refresh: bool = False) -> pd.DataFrame:
    df = _read_sheet(["Cost Codes", "CostCodes"], ["Cost Code", "Cost Code Description", "Active"], force_refresh, path)
    if not df.empty:
        df = df.rename(columns={
            "Cost Code": "cost_code",
            "Cost Code Description": "cost_desc",
        })
    return df


def only_active_cost_codes(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    cols = {c.lower(): c for c in df.columns}
    if "active" in cols:
        col = cols["active"]

        def truthy(value):
            if isinstance(value, bool):
                return value
            s = str(value).strip().lower()
            return s in {"true", "t", "yes", "y", "1", "active", "enabled"}

        return df[df[col].apply(truthy)]
    return df


def pad_job_area(value) -> str:
    s = str(value).strip()
    return f"{int(s):03d}" if s.isdigit() else s


def get_time_data(path: str | None = None, force_refresh: bool = False) -> pd.DataFrame:
    df = _read_sheet(["Time Data", "TimeData"], TIME_DATA_HEADERS, force_refresh, path)
    if df.empty:
        return pd.DataFrame(columns=TIME_DATA_HEADERS)
    for missing in TIME_DATA_HEADERS:
        if missing not in df.columns:
            df[missing] = ""
    return df[TIME_DATA_HEADERS + [c for c in df.columns if c not in TIME_DATA_HEADERS]]


def _ensure_time_data_headers(xlsx_file: Path) -> None:
    wb = load_workbook(xlsx_file)
    if "Time Data" not in wb.sheetnames:
        ws = wb.create_sheet("Time Data")
        ws.append(TIME_DATA_HEADERS)
        wb.save(xlsx_file)
        return
    ws = wb["Time Data"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = [str(h).strip() if h is not None else "" for h in headers]
    changed = False
    for idx, header in enumerate(TIME_DATA_HEADERS, start=1):
        if header not in headers:
            ws.cell(row=1, column=len(headers) + 1, value=header)
            headers.append(header)
            changed = True
    if changed:
        wb.save(xlsx_file)


def append_time_row(path: str | None, payload: dict) -> bool:
    sheet_id = st.secrets.get("google_sheets_id", "")
    manager = get_sheets_manager()
    row = [payload.get(col, "") for col in TIME_DATA_HEADERS]

    if sheet_id:
        try:
            if manager.append_rows("Time Data", [row], sheet_id):
                if hasattr(manager, "_data_cache"):
                    manager._data_cache.pop("Time Data", None)  # type: ignore[attr-defined]
                return True
        except Exception:
            pass

    local_path = _resolve_workbook_path(path)
    if local_path is None:
        st.error("No Google Sheets ID configured and local workbook path not found.")
        return False
    try:
        _ensure_time_data_headers(local_path)
        wb = load_workbook(local_path)
        ws = wb["Time Data"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        headers = [str(h).strip() if h is not None else "" for h in headers]
        if not headers:
            headers = TIME_DATA_HEADERS
            for idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=idx, value=header)
        row_vals = [payload.get(h, "") for h in headers]
        ws.append(row_vals)
        wb.save(local_path)
        return True
    except Exception as exc:
        st.error(f"Failed to append to local workbook: {exc}")
        return False
