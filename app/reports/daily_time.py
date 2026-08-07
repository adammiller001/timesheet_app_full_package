import io
import pandas as pd
from datetime import date
from openpyxl.styles import Font
from app.data.workbook import get_time_data
from app.exports.google_templates import get_google_template_workbook_bytes, load_template_sheet_workbook

def daily_time_report(xlsx_path: str, export_date: date) -> bytes | None:
    td = get_time_data(xlsx_path)
    if td.empty or "Date" not in td.columns:
        return None
    date_str = export_date.strftime("%Y-%m-%d")
    day = td[td["Date"].astype(str).str[:10] == date_str].copy()
    if day.empty:
        return None

    wb, ws = load_template_sheet_workbook(
        get_google_template_workbook_bytes(),
        ("Daily Time", "DailyTime"),
    )

    try:
        ws["B1"] = pd.to_datetime(date_str).strftime("%A, %B %d, %Y")
        ws["B2"] = "2224138065"
        ws["B3"] = "Pembina"
    except Exception:
        pass

    # Work Descriptions
    descs = {}
    if "Comments" in day.columns and "Job Number" in day.columns:
        for job_num in sorted(day["Job Number"].astype(str).str.strip().unique().tolist()):
            job_comments = day[day["Job Number"].astype(str).str.strip() == job_num]["Comments"].dropna()
            if not job_comments.empty:
                texts = job_comments.astype(str).str.strip().replace("nan","").tolist()
                unique_comments, seen = [], set()
                for t in texts:
                    if t and t not in seen:
                        unique_comments.append(t); seen.add(t)
                if unique_comments:
                    descs[str(job_num)] = unique_comments

    bold_underline = Font(bold=True, underline="single")
    for r in range(264, 401):
        ws.cell(row=r, column=1, value=None)
        ws.cell(row=r, column=2, value=None)

    row_ptr = 264
    for job_num, comments in descs.items():
        ws.cell(row=row_ptr, column=1, value="Work Description").font = bold_underline
        ws.cell(row=row_ptr, column=2, value=job_num).font = bold_underline
        row_ptr += 1
        for comment in comments:
            ws.cell(row=row_ptr, column=2, value=comment); row_ptr += 1
        row_ptr += 1

    out = io.BytesIO()
    wb.save(out); out.seek(0)
    return out.getvalue()
